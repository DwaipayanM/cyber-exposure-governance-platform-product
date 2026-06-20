from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from core.asset_context_engine import calculate_business_impact, enrich_with_asset_context
from core.audit_logger import log_event, read_audit_log
from core.control_mapping_engine import add_control_mapping
from core.csv_security import sanitize_df_for_export
from core.exception_engine import apply_exceptions, load_exceptions
from core.ids_correlation_engine import correlate_ids_alerts
from core.import_normalizer import normalize_vulnerability_input
from core.network_exposure_engine import add_network_exposure
from core.playbook_engine import add_remediation_playbooks
from core.policy_engine import load_risk_policy
from core.privacy_impact_engine import add_privacy_impact
from core.remediation_governance import assign_remediation_governance
from core.report_integrity import generate_sha256, verify_sha256
from core.run_history import save_run, list_runs, load_run, delete_run, fingerprint_df, storage_mode
from core.scoring_engine import calculate_final_score, calculate_threat_intelligence_score, summarize_score_drivers
from core.simulation_engine import run_simulation
from core.ticket_exporter import build_ticket_export
from core.trend_engine import read_risk_snapshots, save_risk_snapshot
from core.validator import validate_asset_df, validate_ids_df, validate_vulnerability_df
from services.cisa_kev_service import enrich_with_kev
from services.epss_service import enrich_with_epss
from services.nvd_service import enrich_with_nvd


APP_TITLE = "Cyber Exposure Governance Platform"
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"

# ---------------------------------------------------------------------------
# Brand palette  (blue / grey / green / yellow, red used sparingly for Critical)
# ---------------------------------------------------------------------------
# HPE-inspired accent (brand green + charcoal) layered on the existing layout.
GREEN = "#01A982"        # HPE green - primary brand accent
GREEN_D = "#017A5E"      # darker green, legible as text on white
CHARCOAL = "#1D1F27"     # HPE charcoal - dark header band / headings
NAVY = CHARCOAL          # all former navy surfaces now render as charcoal
BLUE = "#1E50A0"         # retained only for the "Medium" priority semantic
BLUE2 = "#2E6FD0"
GREY = "#5B6675"
GREY_BG = "#EEF1F5"
YELLOW = "#C9A227"
RED = "#B23A3A"
INK = "#1F2937"

PRIORITY_COLORS = {"Critical": RED, "High": YELLOW, "Medium": BLUE, "Low": "#2E7D5B"}

# Pale cell styles for colour-coded tables (readable dark text on tint).
PRIORITY_CELL = {
    "Critical": "background-color:#F4D7D7;color:#8E2A2A;font-weight:600",
    "High": "background-color:#F6ECC9;color:#7A5E10;font-weight:600",
    "Medium": "background-color:#DCE6F6;color:#163A6B;font-weight:600",
    "Low": "background-color:#D6EAE0;color:#1C5740;font-weight:600",
}
SLA_CELL = {
    "Breached": "background-color:#F4D7D7;color:#8E2A2A;font-weight:600",
    "Due Soon": "background-color:#F6ECC9;color:#7A5E10;font-weight:600",
    "Within SLA": "background-color:#D6EAE0;color:#1C5740;font-weight:600",
}
YESNO_CELL = {
    "Yes": "background-color:#F4D7D7;color:#8E2A2A;font-weight:600",
}

st.set_page_config(page_title=APP_TITLE, page_icon="\U0001F6E1\uFE0F", layout="wide")


# ---------------------------------------------------------------------------
# Theme / look-and-feel
# ---------------------------------------------------------------------------
def inject_theme() -> None:
    st.markdown(
        """
        <style>
        .block-container {padding-top: 1.4rem; padding-bottom: 2rem; max-width: 1300px;}
        html, body, [class*="css"] {color: #1F2937;}

        /* ---- Header banner ---- */
        .cegp-header {
            background: linear-gradient(115deg, #1D1F27 0%, #23262F 52%, #015E48 88%, #01A982 122%);
            border-radius: 14px; padding: 20px 26px; color: #fff; margin-bottom: 14px;
            box-shadow: 0 6px 18px rgba(21,49,92,.18);
        }
        .cegp-header-row {display:flex; align-items:center; gap:16px;}
        .cegp-logo {font-size: 2.1rem; line-height:1; display:flex; align-items:center;}
        .cegp-authors {font-size:.82rem; color:#EAF6F1; margin-top:3px; font-weight:600;}
        .cegp-conceptcard {background:#F4F8F6; border:1px solid #CDE8DF; border-left:4px solid #01A982;
            border-radius:8px; padding:10px 14px; font-size:.9rem; color:#15315C; margin:6px 0 10px;}
        .cegp-chip {cursor: help;}
        .cegp-title {font-size: 1.5rem; font-weight: 800; letter-spacing:.2px;}
        .cegp-tag {font-size: .9rem; opacity:.92; margin-top:2px;}
        .cegp-badge {
            margin-left:auto; background: rgba(255,255,255,.16); border:1px solid rgba(255,255,255,.35);
            padding: 4px 12px; border-radius: 999px; font-size:.78rem; font-weight:700; letter-spacing:.05em;
        }
        .cegp-chips {margin-top:12px; display:flex; flex-wrap:wrap; gap:8px;}
        .cegp-chip {
            background: rgba(255,255,255,.14); border:1px solid rgba(255,255,255,.28);
            padding: 3px 11px; border-radius: 999px; font-size:.74rem; font-weight:600;
        }

        /* ---- KPI cards ---- */
        .cegp-kpi {
            background:#fff; border:1px solid #E3E8EF; border-radius:12px; padding:13px 16px;
            box-shadow:0 1px 3px rgba(20,49,92,.06); min-height:92px;
        }
        .cegp-kpi-label {font-size:.70rem; letter-spacing:.05em; text-transform:uppercase; color:#5B6675; font-weight:700;}
        .cegp-kpi-value {font-size:1.85rem; font-weight:800; line-height:1.1; margin-top:6px;}
        .cegp-kpi-sub {font-size:.72rem; color:#8A94A6; margin-top:3px;}

        /* ---- Section headers ---- */
        .cegp-section {
            padding:7px 0 7px 12px; margin:14px 0 8px; font-size:1.05rem; font-weight:700; color:#1D1F27;
            background:#F7F9FC; border-radius:0 8px 8px 0;
        }

        /* ---- Tabs: dark band matching the header gradient, green active pill ---- */
        .stTabs [data-baseweb="tab-list"] {
            gap:4px; margin-top:10px; padding:7px; border-bottom:none; flex-wrap:wrap;
            background:linear-gradient(115deg,#1D1F27 0%,#23262F 52%,#015E48 88%,#01A982 122%);
            border-radius:12px; box-shadow:0 4px 14px rgba(21,49,92,.18);
        }
        .stTabs [data-baseweb="tab"] {
            padding:9px 16px; border-radius:8px; color:#D7DEE8; font-weight:700; background:transparent;
        }
        .stTabs [data-baseweb="tab"]:hover {background:rgba(255,255,255,.10); color:#ffffff;}
        .stTabs [data-baseweb="tab"] p {font-weight:700 !important; font-size:.95rem !important;}
        .stTabs [aria-selected="true"] {background:#01A982 !important; color:#ffffff !important;}
        .stTabs [aria-selected="true"] p {color:#ffffff !important;}
        .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] {display:none;}

        /* ---- Buttons ---- */
        .stButton>button, .stDownloadButton>button {border-radius:8px; font-weight:600;}

        /* ---- Metric (fallback) ---- */
        [data-testid="stMetric"] {
            background:#fff; border:1px solid #E3E8EF; border-radius:12px; padding:10px 14px;
        }

        /* ---- Sidebar ---- */
        [data-testid="stSidebar"] {
            background:linear-gradient(180deg,#F6F9FC 0%, #EEF4F1 100%);
            border-right:1px solid #E3E8EF;
        }
        /* pull sidebar options up to remove the top gap (collapse the empty header band) */
        [data-testid="stSidebarHeader"] {padding-top:0.3rem !important; padding-bottom:0 !important; min-height:0 !important; height:auto !important;}
        [data-testid="stSidebarUserContent"] {padding-top:0.4rem !important;}
        [data-testid="stSidebarContent"] {padding-top:0 !important;}
        [data-testid="stSidebar"] > div:first-child {padding-top:0 !important;}
        [data-testid="stSidebar"] .block-container {padding-top:0.4rem !important;}
        [data-testid="stSidebar"] h2 {color:#15315C;}
        /* Brand banner at top of sidebar */
        .cegp-sidebrand {
            display:flex; align-items:center; gap:10px;
            background:linear-gradient(120deg,#15171E 0%,#1D1F27 55%,#015E48 130%);
            color:#fff; border-radius:12px; padding:12px 14px; margin-bottom:12px;
            box-shadow:0 4px 12px rgba(21,49,92,.18);
        }
        .cegp-sidebrand .b1 {font-weight:800; font-size:1rem; line-height:1.1;}
        .cegp-sidebrand .b2 {font-size:.72rem; color:#CFE9DF; margin-top:2px;}
        /* Section label */
        .cegp-sidelabel {
            font-size:.74rem; font-weight:800; letter-spacing:.06em; text-transform:uppercase;
            color:#15315C; border-left:4px solid #01A982; padding:4px 0 4px 10px; margin:6px 0 10px;
            background:#FFFFFF; border-radius:0 6px 6px 0;
        }
        /* Expanders */
        [data-testid="stSidebar"] details {
            background:#FFFFFF; border:1px solid #E3E8EF; border-radius:10px; margin-bottom:8px;
        }
        [data-testid="stSidebar"] details summary {font-weight:700; color:#15315C;}
        [data-testid="stSidebar"] details summary:hover {color:#017A5E;}
        /* Inputs */
        [data-testid="stSidebar"] [data-testid="stTextInput"] input,
        [data-testid="stSidebar"] [data-baseweb="select"] > div {border-radius:8px;}
        [data-testid="stSidebar"] button[data-testid="stBaseButton-secondary"] {
            font-family: 'Outfit', 'Inter', 'Segoe UI', sans-serif !important;
            font-weight: 800 !important;
            font-size: 1.05rem !important;
            color: #ffffff !important;
            border: 2px solid #01A982 !important;
            border-radius: 8px !important;
            background-color: #01A982 !important;
            box-shadow: 0 2px 6px rgba(1, 169, 130, 0.25) !important;
        }
        [data-testid="stSidebar"] button[data-testid="stBaseButton-secondary"] p {color:#ffffff !important;}
        [data-testid="stSidebar"] button[data-testid="stBaseButton-secondary"]:hover {
            color: #ffffff !important;
            background-color: #017A5E !important;
            border-color: #017A5E !important;
            box-shadow: 0 4px 12px rgba(1, 122, 94, 0.30) !important;
        }

        /* ---- Dataframe ---- */
        [data-testid="stDataFrame"] {border:1px solid #E3E8EF; border-radius:10px;}
        </style>
        """,
        unsafe_allow_html=True,
    )


CYBER_LOGO_SVG = (
    '<svg width="50" height="50" viewBox="0 0 64 64" fill="none" xmlns="http://www.w3.org/2000/svg">'
    '<path d="M32 3 L56 12 V30 C56 46 46 56 32 61 C18 56 8 46 8 30 V12 Z" '
    'fill="#01A982" stroke="#ffffff" stroke-width="2.5" stroke-linejoin="round"/>'
    '<rect x="22" y="29" width="20" height="16" rx="2.5" fill="#15315C" stroke="#ffffff" stroke-width="2"/>'
    '<path d="M26 29 V24 a6 6 0 0 1 12 0 V29" fill="none" stroke="#ffffff" stroke-width="2.5"/>'
    '<circle cx="32" cy="36" r="2.4" fill="#ffffff"/><rect x="31" y="37" width="2" height="5" rx="1" fill="#ffffff"/>'
    '</svg>'
)

# Concept definitions for the sidebar guide (platform building blocks only).
CONCEPTS = {
    "CVSS": "Common Vulnerability Scoring System \u2014 an industry-standard 0\u201310 score for a vulnerability's technical severity (NVD).",
    "EPSS": "Exploit Prediction Scoring System (FIRST) \u2014 probability (0\u20131) that a CVE will be exploited in the wild within 30 days.",
    "CISA KEV": "CISA Known Exploited Vulnerabilities \u2014 an authoritative catalog of CVEs with confirmed real-world exploitation.",
    "IDS/IPS": "Intrusion Detection / Prevention System \u2014 sensors that detect or block malicious traffic; alerts are correlated to vulnerable assets here.",
    "Privacy Impact": "Assessment of exposure to personal/sensitive data \u2014 PII, sensitivity, encryption status and regulatory impact.",
    "SLA Governance": "Remediation Service-Level governance \u2014 priority-based due dates with breach and escalation tracking.",
    "Exposure Score": "The platform's weighted 0\u2013100 cyber-exposure score combining threat intel, business, network, IDS and privacy.",
}
DIAGRAM_CONCEPTS = ["CVSS", "EPSS", "CISA KEV", "IDS/IPS", "Privacy Impact", "SLA Governance", "Exposure Score"]

# Fuller, plain-English explanations shown in the sidebar concept guide.
CONCEPT_DETAILS = {
    "CVSS": "**Common Vulnerability Scoring System.** An industry-standard score from **0 to 10** rating a vulnerability's *technical severity* \u2014 how damaging it could be if exploited \u2014 published in the NVD. Bands: Low (0.1\u20133.9), Medium (4\u20136.9), High (7\u20138.9), Critical (9\u201310). It measures impact only, **not** whether anyone is actually exploiting it, so the platform combines it with real-world threat and business context.",
    "EPSS": "**Exploit Prediction Scoring System** (by FIRST). A **probability (0\u20131)** that a CVE will be exploited in the wild within the next **30 days** \u2014 a high value means attackers are likely to use it soon. The platform uses the EPSS percentile to weight *likelihood* alongside severity, so vulnerabilities attackers actually use rise to the top.",
    "CISA KEV": "**Known Exploited Vulnerabilities** \u2014 the U.S. CISA catalogue of CVEs **confirmed to be exploited in real attacks**. If a CVE is on this list, exploitation is not theoretical, it is happening now. The platform treats KEV membership as a strong escalator of priority.",
    "IDS/IPS": "**Intrusion Detection / Prevention Systems** \u2014 network sensors that *detect* (IDS) or *block* (IPS) malicious traffic. Their alerts are correlated to your vulnerable assets, so a host that is **both vulnerable and currently being targeted** is flagged for urgent action \u2014 the live \u201cis it under attack now?\u201d signal scanners miss.",
    "Privacy Impact": "**How much sensitive or personal data an asset holds, and how exposed it is.** The platform weighs **PII** presence, data sensitivity, **encryption status**, and regulatory impact (e.g. GDPR / HIPAA). An unencrypted system holding regulated personal data raises both the exposure score and the compliance stakes.",
    "SLA Governance": "**Service-Level governance for remediation.** Each exposure gets a priority-based **due date** (Critical = 7 days, High = 15, Medium = 30, Low = 90) and is tracked as **Within SLA / Due Soon / Breached**, with escalations for overdue items. This turns \u201cwe should fix this\u201d into an accountable, time-bound commitment.",
    "Exposure Score": "**The platform's headline number** \u2014 a single, **explainable 0\u2013100 score** combining threat intelligence (35%), network exposure (20%), business impact (15%), IDS correlation (15%), privacy impact (10%) and SLA governance (5%). It maps to a Low / Medium / High / Critical priority, and because every point traces back to a factor, the result is **defensible, not a black box**.",
}


def render_header() -> None:
    st.markdown(
        f"""
        <div class="cegp-header">
          <div class="cegp-header-row">
            <span class="cegp-logo">{CYBER_LOGO_SVG}</span>
            <div>
              <div class="cegp-title">Cyber Exposure Governance Platform (CEGP)</div>
              <div class="cegp-authors">Dwaipayan Mojumder &middot; Deblina Das &nbsp;|&nbsp; M.Sc. Cyber Security (4th Sem) &nbsp;|&nbsp; Under the guidance of Prof. Sanjay Pal</div>
              <div class="cegp-tag">Risk-based vulnerability prioritization &middot; remediation governance &middot; audit-ready evidence</div>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _diagram_svg(concept: str) -> str:
    """Small palette-consistent SVG explainer for a platform concept."""
    G, C, Y, R, GR = "#01A982", "#15315C", "#C9A227", "#B23A3A", "#5B6675"
    base = 'font-family="Segoe UI,Roboto,sans-serif"'
    if concept == "CVSS":
        return (f'<svg viewBox="0 0 460 120" {base} xmlns="http://www.w3.org/2000/svg">'
                f'<text x="8" y="20" font-size="13" fill="{C}" font-weight="700">CVSS severity scale (0\u201310)</text>'
                f'<rect x="8" y="38" width="100" height="26" fill="#D6EAE0"/><rect x="108" y="38" width="120" height="26" fill="#DCE6F6"/>'
                f'<rect x="228" y="38" width="120" height="26" fill="#F6ECC9"/><rect x="348" y="38" width="104" height="26" fill="#F4D7D7"/>'
                f'<text x="58" y="56" font-size="11" fill="#1C5740" text-anchor="middle">Low 0.1\u20133.9</text>'
                f'<text x="168" y="56" font-size="11" fill="#163A6B" text-anchor="middle">Medium 4\u20136.9</text>'
                f'<text x="288" y="56" font-size="11" fill="#7A5E10" text-anchor="middle">High 7\u20138.9</text>'
                f'<text x="400" y="56" font-size="11" fill="#8E2A2A" text-anchor="middle">Critical 9\u201310</text>'
                f'<text x="8" y="92" font-size="11" fill="{GR}">Technical impact only \u2014 the platform combines it with exploitation &amp; business context.</text></svg>')
    if concept == "EPSS":
        return (f'<svg viewBox="0 0 460 120" {base} xmlns="http://www.w3.org/2000/svg">'
                f'<text x="8" y="20" font-size="13" fill="{C}" font-weight="700">EPSS exploitation probability (percentile)</text>'
                f'<rect x="8" y="36" width="440" height="14" rx="7" fill="#EEF1F5"/>'
                f'<rect x="8" y="36" width="352" height="14" rx="7" fill="{G}"/>'
                f'<line x1="228" y1="32" x2="228" y2="54" stroke="{Y}" stroke-width="2"/><text x="228" y="68" font-size="10" fill="{Y}" text-anchor="middle">0.50</text>'
                f'<line x1="360" y1="32" x2="360" y2="54" stroke="{R}" stroke-width="2"/><text x="360" y="68" font-size="10" fill="{R}" text-anchor="middle">0.80</text>'
                f'<text x="8" y="96" font-size="11" fill="{GR}">Higher percentile = more likely to be exploited soon \u2192 higher threat score.</text></svg>')
    if concept == "CISA KEV":
        return (f'<svg viewBox="0 0 460 120" {base} xmlns="http://www.w3.org/2000/svg">'
                f'<rect x="8" y="34" width="120" height="40" rx="6" fill="#EEF1F5" stroke="{GR}"/><text x="68" y="58" font-size="11" fill="{C}" text-anchor="middle">CVE in KEV?</text>'
                f'<path d="M128 54 H180" stroke="{GR}" stroke-width="2" marker-end="url(#a)"/>'
                f'<rect x="180" y="34" width="120" height="40" rx="6" fill="#F4D7D7" stroke="{R}"/><text x="240" y="58" font-size="11" fill="#8E2A2A" text-anchor="middle">Confirmed exploited</text>'
                f'<path d="M300 54 H352" stroke="{GR}" stroke-width="2" marker-end="url(#a)"/>'
                f'<rect x="352" y="34" width="100" height="40" rx="6" fill="{G}"/><text x="402" y="58" font-size="11" fill="#fff" text-anchor="middle">Prioritise</text>'
                f'<defs><marker id="a" markerWidth="8" markerHeight="8" refX="6" refY="3" orient="auto"><path d="M0 0 L6 3 L0 6" fill="{GR}"/></marker></defs>'
                f'<text x="8" y="100" font-size="11" fill="{GR}">KEV membership is a strong "fix this now" signal.</text></svg>')
    if concept == "IDS/IPS":
        return (f'<svg viewBox="0 0 460 120" {base} xmlns="http://www.w3.org/2000/svg">'
                f'<circle cx="40" cy="54" r="16" fill="#F4D7D7" stroke="{R}"/><text x="40" y="58" font-size="10" fill="#8E2A2A" text-anchor="middle">Attacker</text>'
                f'<path d="M58 54 H120" stroke="{GR}" stroke-width="2" marker-end="url(#b)"/>'
                f'<rect x="120" y="40" width="80" height="28" rx="6" fill="{C}"/><text x="160" y="58" font-size="10" fill="#fff" text-anchor="middle">IDS/IPS</text>'
                f'<path d="M200 54 H262" stroke="{GR}" stroke-width="2" marker-end="url(#b)"/>'
                f'<rect x="262" y="40" width="86" height="28" rx="6" fill="{Y}"/><text x="305" y="58" font-size="10" fill="#5b4708" text-anchor="middle">Alert</text>'
                f'<path d="M348 54 H410" stroke="{GR}" stroke-width="2" marker-end="url(#b)"/>'
                f'<rect x="410" y="40" width="44" height="28" rx="6" fill="{G}"/><text x="432" y="58" font-size="9" fill="#fff" text-anchor="middle">Asset</text>'
                f'<defs><marker id="b" markerWidth="8" markerHeight="8" refX="6" refY="3" orient="auto"><path d="M0 0 L6 3 L0 6" fill="{GR}"/></marker></defs>'
                f'<text x="8" y="100" font-size="11" fill="{GR}">Live alerts are correlated to the vulnerable asset to confirm active targeting.</text></svg>')
    if concept == "Privacy Impact":
        return (f'<svg viewBox="0 0 460 120" {base} xmlns="http://www.w3.org/2000/svg">'
                f'<rect x="8" y="34" width="84" height="40" rx="6" fill="#EEF1F5" stroke="{GR}"/><text x="50" y="58" font-size="11" fill="{C}" text-anchor="middle">PII?</text>'
                f'<rect x="118" y="34" width="110" height="40" rx="6" fill="#EEF1F5" stroke="{GR}"/><text x="173" y="53" font-size="10" fill="{C}" text-anchor="middle">Sensitivity /</text><text x="173" y="66" font-size="10" fill="{C}" text-anchor="middle">Encryption</text>'
                f'<rect x="254" y="34" width="110" height="40" rx="6" fill="#EEF1F5" stroke="{GR}"/><text x="309" y="58" font-size="10" fill="{C}" text-anchor="middle">Regulatory</text>'
                f'<rect x="390" y="34" width="62" height="40" rx="6" fill="{R}"/><text x="421" y="58" font-size="10" fill="#fff" text-anchor="middle">Impact</text>'
                f'<text x="8" y="100" font-size="11" fill="{GR}">Unencrypted, regulated personal data raises the privacy dimension of the score.</text></svg>')
    if concept == "SLA Governance":
        return (f'<svg viewBox="0 0 460 120" {base} xmlns="http://www.w3.org/2000/svg">'
                f'<text x="8" y="20" font-size="13" fill="{C}" font-weight="700">Detected \u2192 SLA window \u2192 status</text>'
                f'<circle cx="20" cy="50" r="6" fill="{C}"/><text x="14" y="74" font-size="10" fill="{GR}">Detected</text>'
                f'<rect x="20" y="46" width="300" height="8" fill="#D6EAE0"/><rect x="320" y="46" width="120" height="8" fill="#F4D7D7"/>'
                f'<line x1="320" y1="40" x2="320" y2="60" stroke="{C}" stroke-width="2"/><text x="320" y="74" font-size="10" fill="{C}" text-anchor="middle">Due date</text>'
                f'<text x="160" y="40" font-size="10" fill="#1C5740" text-anchor="middle">Within SLA</text><text x="380" y="40" font-size="10" fill="#8E2A2A" text-anchor="middle">Breached</text>'
                f'<text x="8" y="100" font-size="11" fill="{GR}">Due dates derive from priority (Critical 7d, High 15d, Medium 30d, Low 90d).</text></svg>')
    # Exposure Score
    return (f'<svg viewBox="0 0 460 130" {base} xmlns="http://www.w3.org/2000/svg">'
            f'<text x="8" y="18" font-size="13" fill="{C}" font-weight="700">Weighted exposure score (0\u2013100)</text>'
            + "".join(
                f'<rect x="8" y="{30+i*14}" width="{w*3}" height="10" rx="3" fill="{col}"/>'
                f'<text x="{14+w*3}" y="{39+i*14}" font-size="10" fill="{GR}">{lbl} ({w})</text>'
                for i,(lbl,w,col) in enumerate([("Threat intel",35,G),("Network",20,C),("Business",15,Y),("IDS",15,R),("Privacy",10,"#2E7D5B"),("SLA",5,GR)]))
            + f'<text x="8" y="126" font-size="11" fill="{GR}">Components sum to a single, explainable 0\u2013100 priority score.</text></svg>')


def render_sidebar_concepts() -> None:
    """Tucked-away concept guide at the bottom of the sidebar (only if expanded)."""
    with st.expander("\u2139\ufe0f  Concept guide", expanded=False):
        st.caption("What each metric means and how the platform uses it.")
        choice = st.selectbox(
            "View a concept", ["\u2014 all concepts \u2014"] + DIAGRAM_CONCEPTS,
            index=0, key="concept_pick", label_visibility="collapsed",
        )
        if choice and choice != "\u2014 all concepts \u2014":
            st.markdown(_diagram_svg(choice), unsafe_allow_html=True)
            st.markdown(CONCEPT_DETAILS.get(choice, CONCEPTS.get(choice, "")))
        else:
            for k in DIAGRAM_CONCEPTS:
                st.markdown(f"#### {k}")
                st.markdown(CONCEPT_DETAILS.get(k, CONCEPTS.get(k, "")))
                st.divider()


@st.dialog("Project Documentation", width="large")
def show_documentation() -> None:
    doc_path = BASE_DIR / "docs" / "project_documentation.md"
    try:
        st.markdown(doc_path.read_text(encoding="utf-8"))
    except Exception:
        st.warning("Documentation file not found at docs/project_documentation.md.")
    try:
        st.download_button("Download documentation (.md)", data=doc_path.read_bytes(),
                           file_name="project_documentation.md", mime="text/markdown")
    except Exception:
        pass


@st.dialog("Future Scope — Implementation using GCP", width="large")
def show_gcp_guide() -> None:
    guide_path = BASE_DIR / "docs" / "gcp_deployment_guide.md"
    try:
        st.markdown(guide_path.read_text(encoding="utf-8"))
    except Exception:
        st.warning("Guide not found at docs/gcp_deployment_guide.md.")
    try:
        st.download_button("Download GCP guide (.md)", data=guide_path.read_bytes(),
                           file_name="gcp_deployment_guide.md", mime="text/markdown")
    except Exception:
        pass


def section(title: str, accent: str = GREEN) -> None:
    st.markdown(
        f'<div class="cegp-section" style="border-left:4px solid {accent}">{title}</div>',
        unsafe_allow_html=True,
    )


def kpi_card(col, label: str, value, border: str, value_color: str, sub: str = "") -> None:
    col.markdown(
        f"""
        <div class="cegp-kpi" style="border-top:4px solid {border}">
          <div class="cegp-kpi-label">{label}</div>
          <div class="cegp-kpi-value" style="color:{value_color}">{value}</div>
          <div class="cegp-kpi-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def style_fig(fig, color: str = GREEN):
    fig.update_layout(
        template="plotly_white",
        font=dict(family="Segoe UI, Roboto, sans-serif", color=INK, size=13),
        title=dict(font=dict(size=15, color=NAVY), x=0.01, xanchor="left"),
        margin=dict(l=10, r=10, t=54, b=48),
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        legend=dict(orientation="h", yanchor="top", y=-0.18, x=0, title_text=""),
    )
    fig.update_xaxes(showgrid=False, title_font=dict(color=GREY), tickfont=dict(color=GREY))
    fig.update_yaxes(gridcolor="#EDF1F6", title_font=dict(color=GREY), tickfont=dict(color=GREY))
    if not fig.data or all(getattr(t, "marker", None) is None for t in fig.data):
        return fig
    for tr in fig.data:
        marker = getattr(tr, "marker", None)
        if marker is None:
            continue
        # Pie/donut markers use .colors (plural); only set .color where it exists and is unset.
        if "color" in getattr(marker, "_props", {}) or hasattr(type(marker), "color"):
            try:
                if marker.color is None:
                    marker.color = color
            except (AttributeError, ValueError):
                pass
    return fig


def _style_col(styler, col, mapping):
    if col not in styler.data.columns:
        return styler
    fn = getattr(styler, "map", None) or styler.applymap
    return fn(lambda v: mapping.get(str(v), ""), subset=[col])


def styled(df: pd.DataFrame):
    sty = df.style
    sty = _style_col(sty, "priority", PRIORITY_CELL)
    sty = _style_col(sty, "sla_status", SLA_CELL)
    sty = _style_col(sty, "kev_status", YESNO_CELL)
    sty = _style_col(sty, "exploit_attempt_detected", YESNO_CELL)
    return sty


_ACRONYMS = {"cve", "id", "kev", "ids", "ips", "sla", "epss", "cvss", "pii", "ip", "vpn", "hmac", "os", "tco", "url"}


def titleize(col: str) -> str:
    """Human, init-caps column header with security acronyms kept uppercase."""
    words = str(col).replace("_", " ").split()
    return " ".join(w.upper() if w.lower() in _ACRONYMS else w.capitalize() for w in words)


# Header + body styling for HTML tables (palette-consistent: navy header, green underline).
_TABLE_STYLES = [
    # let columns size to their content instead of being squeezed into 100% width
    {"selector": "", "props": [("border-collapse", "collapse"), ("width", "auto"), ("min-width", "100%"),
                               ("font-family", "Segoe UI, Roboto, sans-serif")]},
    {"selector": "th.col_heading",
     "props": [("background-color", "#15315C"), ("color", "#FFFFFF"), ("font-weight", "700"),
               ("text-align", "left"), ("padding", "8px 12px"), ("font-size", ".82rem"),
               ("border-bottom", "2px solid #01A982"), ("position", "sticky"), ("top", "0"),
               ("z-index", "2"), ("white-space", "nowrap"), ("resize", "horizontal"), ("overflow", "hidden")]},
    {"selector": "td", "props": [("padding", "6px 12px"), ("border-top", "1px solid #EDF1F6"),
                                  ("font-size", ".84rem"), ("color", "#1F2937"), ("white-space", "nowrap"),
                                  ("max-width", "360px"), ("overflow", "hidden"), ("text-overflow", "ellipsis"),
                                  ("vertical-align", "middle")]},
    {"selector": "tbody tr:nth-child(even) td", "props": [("background-color", "#FBFCFE")]},
    {"selector": "tbody tr:hover td", "props": [("background-color", "#F1F7F4")]},
]


def render_table(df: pd.DataFrame, height: int = 430) -> None:
    """Render a dataframe as a styled HTML table with bold, init-caps, coloured headers."""
    d = df.copy()
    d.columns = [titleize(c) for c in d.columns]
    sty = d.style
    sty = _style_col(sty, "Priority", PRIORITY_CELL)
    sty = _style_col(sty, "SLA Status", SLA_CELL)
    sty = _style_col(sty, "KEV Status", YESNO_CELL)
    sty = _style_col(sty, "Exploit Attempt Detected", YESNO_CELL)
    try:
        sty = sty.hide(axis="index")
    except Exception:
        pass
    sty = sty.set_table_styles(_TABLE_STYLES)
    st.markdown(
        f'<div style="max-height:{height}px;overflow:auto;border:1px solid #E3E8EF;border-radius:10px">{sty.to_html()}</div>',
        unsafe_allow_html=True,
    )


def operator() -> str:
    return st.session_state.get("operator", "system") or "system"


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------
def load_csv(uploaded_file, default_path: Path) -> pd.DataFrame:
    """Load an uploaded CSV or Excel file (.csv/.xlsx/.xls), else the bundled CSV.

    The Excel engine is chosen by extension: openpyxl for .xlsx, xlrd for legacy .xls.
    """
    if uploaded_file is not None:
        name = str(getattr(uploaded_file, "name", "")).lower()
        if name.endswith(".xlsx"):
            return pd.read_excel(uploaded_file, engine="openpyxl")  # first sheet
        if name.endswith(".xls"):
            return pd.read_excel(uploaded_file, engine="xlrd")      # legacy Excel
        return pd.read_csv(uploaded_file)
    return pd.read_csv(default_path)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    # Sanitise against CSV / formula injection before producing the file bytes.
    return sanitize_df_for_export(df).to_csv(index=False).encode("utf-8")


def executive_summary_report(df: pd.DataFrame) -> pd.DataFrame:
    summary_rows = [
        {"Metric": "Total Exposures", "Value": len(df)},
        {"Metric": "Critical Exposures", "Value": int((df["priority"] == "Critical").sum())},
        {"Metric": "High Exposures", "Value": int((df["priority"] == "High").sum())},
        {"Metric": "CISA KEV Exposures", "Value": int((df["kev_status"] == "Yes").sum())},
        {"Metric": "IDS-Correlated Exposures", "Value": int((df["ids_alert_count"].fillna(0).astype(int) > 0).sum())},
        {"Metric": "Privacy Critical/High Exposures", "Value": int(df["privacy_impact_level"].isin(["Critical", "High"]).sum())},
        {"Metric": "SLA Breached Items", "Value": int((df["sla_status"] == "Breached").sum())},
        {"Metric": "Average Risk Score", "Value": round(float(df["final_score"].mean()), 2) if len(df) else 0},
    ]
    return pd.DataFrame(summary_rows)


# ---------------------------------------------------------------------------
# Remediation Action Plan — compose readable, per-exposure guidance from the
# columns the engines already produce (+ NVD description) and export it.
# ---------------------------------------------------------------------------

# Column-name candidates (the NVD enricher may name these differently across builds).
_DESC_COLS = ["cve_description", "nvd_description", "description", "vuln_description", "summary", "cve_summary"]
_CWE_COLS = ["cwe", "cwe_id", "cwe_name", "weakness", "cwe_description"]


def _val(row, names, default: str = "") -> str:
    """First present, non-empty value among candidate column names, as a clean string."""
    if isinstance(names, str):
        names = [names]
    for n in names:
        if n in row.index:
            v = row[n]
            if pd.notna(v) and str(v).strip().lower() not in ("", "nan", "none"):
                return str(v).strip()
    return default


def _is_yes(row, col) -> bool:
    return _val(row, col).strip().lower() in ("yes", "true", "1")


def _ids_active(row) -> int:
    try:
        return int(float(_val(row, "ids_alert_count", "0")))
    except Exception:
        return 0


def explain_vulnerability(row) -> str:
    """Plain-language description of what the vulnerability is."""
    cve = _val(row, "cve_id", "This exposure")
    asset = _val(row, "asset_name", "the affected asset")
    app = _val(row, "application_name")
    sev, cvss = _val(row, "severity"), _val(row, "cvss_score")
    desc, cwe = _val(row, _DESC_COLS), _val(row, _CWE_COLS)

    meta = ", ".join([m for m in (f"{sev} severity" if sev else "", f"CVSS {cvss}" if cvss else "") if m])
    lead = f"{cve}" + (f" ({meta})" if meta else "")
    target = asset + (f" — application '{app}'" if app else "")
    out = [f"{lead} affects {target}."]
    if desc:
        out.append(desc if desc.endswith(".") else desc + ".")
    else:
        out.append("It is a known security weakness in the affected component that an attacker could "
                   "leverage to compromise the asset's confidentiality, integrity, or availability.")
    if cwe:
        out.append(f"Weakness class: {cwe}.")
    return " ".join(out)


def possible_impact(row) -> str:
    """Why it matters — business, exploitation, network, IDS and privacy consequences."""
    bits = []
    bp = _val(row, "business_process")
    if bp:
        bits.append(f"The affected asset supports the {bp} business process, so a compromise has direct business consequences.")
    if _is_yes(row, "kev_status"):
        bits.append("This CVE is on the CISA KEV list — confirmed exploited in real-world attacks, so the threat is active, not theoretical.")
    nel, ner = _val(row, "network_exposure_level"), _val(row, "network_exposure_reason")
    if nel:
        bits.append(f"Network exposure is {nel}" + (f" ({ner})." if ner else "."))
    n = _ids_active(row)
    if n > 0:
        sevh = _val(row, "highest_alert_severity")
        bits.append(f"IDS/IPS has correlated {n} live alert(s)" + (f" (highest severity: {sevh})" if sevh else "")
                    + " to this asset, indicating possible active targeting.")
    pil, reg = _val(row, "privacy_impact_level"), _val(row, "regulatory_impact")
    if pil and pil.lower() not in ("none", "low"):
        bits.append(f"Privacy impact is {pil}" + (f", carrying regulatory exposure ({reg})." if reg else "."))
    drivers = _val(row, "score_drivers")
    if drivers:
        bits.append(f"Key risk drivers: {drivers}.")
    if not bits:
        bits.append("If left unremediated, this exposure raises the asset's overall cyber risk and could be chained into a broader compromise.")
    return " ".join(bits)


def remediation_steps(row) -> list:
    """Ordered list of concrete fix steps."""
    steps = []
    pa = _val(row, "primary_action")
    if pa:
        steps.append(pa)
    if _is_yes(row, "kev_status"):
        steps.append("Treat as urgent and schedule ahead of non-KEV items, given confirmed in-the-wild exploitation.")
    vs = _val(row, "validation_step")
    if vs:
        steps.append(f"Validate the fix — {vs}")
    ca = _val(row, "control_area")
    if ca:
        steps.append(f"Confirm the related security control is restored or strengthened ({ca}).")
    if not steps:
        steps.append("Apply the vendor-supplied patch or upgrade to a fixed version, then re-scan to confirm the CVE is no longer detected.")
    return steps


def interim_mitigation(row) -> list:
    """What to do now if a full fix is not yet possible."""
    bits = []
    tm = _val(row, "temporary_mitigation")
    cc = _val(row, "compensating_control")
    if tm:
        bits.append(tm)
    if cc:
        bits.append(f"Compensating control — {cc}")
    if not bits:
        bits.append("If immediate patching is not possible, restrict network access to the asset, increase monitoring, "
                    "and limit privileged access until the fix is applied.")
    return bits


def compose_action_plan(df_in: pd.DataFrame) -> pd.DataFrame:
    """Flatten per-exposure guidance into one export-ready table (single source for cards + files)."""
    rows = []
    for _, r in df_in.iterrows():
        rows.append({
            "Rank": len(rows) + 1,
            "Asset": _val(r, "asset_name"),
            "Application": _val(r, "application_name"),
            "CVE": _val(r, "cve_id"),
            "Priority": _val(r, "priority"),
            "Exposure Score": _val(r, "final_score"),
            "KEV": _val(r, "kev_status"),
            "SLA Status": _val(r, "sla_status"),
            "Owner": _val(r, "asset_owner"),
            "Due Date": _val(r, "remediation_due_date"),
            "Vulnerability Explanation": explain_vulnerability(r),
            "Possible Impact": possible_impact(r),
            "Remediation Steps": "\n".join(f"{i+1}. {s}" for i, s in enumerate(remediation_steps(r))),
            "Interim Mitigation": "\n".join(f"- {s}" for s in interim_mitigation(r)),
            "Control Area": _val(r, "control_area"),
            "Escalation": _val(r, "escalation_reason") or _val(r, "escalation_required"),
        })
    return pd.DataFrame(rows)


def to_xlsx_bytes(df_in: pd.DataFrame, sheet_name: str = "Action Plan") -> bytes:
    """Export a sanitised, formatted .xlsx (wrapped text, branded header, frozen header row)."""
    from openpyxl.styles import Font, PatternFill, Alignment

    d = sanitize_df_for_export(df_in)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        d.to_excel(xl, index=False, sheet_name=sheet_name)
        ws = xl.sheets[sheet_name]
        head_fill = PatternFill("solid", fgColor="15315C")
        head_font = Font(color="FFFFFF", bold=True)
        wrap_top = Alignment(wrap_text=True, vertical="top")
        wide = {"Vulnerability Explanation", "Possible Impact", "Remediation Steps", "Interim Mitigation"}
        for ci, col in enumerate(d.columns, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.fill, cell.font, cell.alignment = head_fill, head_font, Alignment(vertical="center")
            ws.column_dimensions[cell.column_letter].width = 60 if col in wide else max(12, min(30, len(str(col)) + 4))
        for r_ in ws.iter_rows(min_row=2):
            for cell in r_:
                cell.alignment = wrap_top
        ws.freeze_panes = "A2"
    return buf.getvalue()


def to_pdf_bytes(plan_df: pd.DataFrame, generated_by: str = "system") -> bytes:
    """Export the action plan as a consulting-grade PDF.

    Follows the house document standard: 0.5 cm margins on all sides, justified
    body text, no header/footer except a page number at bottom-right, fixed-width
    boxes/tables that always stay within the usable page width (never overflow),
    and a formal blue / grey / green / yellow palette (red used sparingly).
    """
    from datetime import datetime
    from xml.sax.saxutils import escape
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether)

    # Palette (matches the app's PRIORITY_COLORS: Critical red, High yellow, Medium blue, Low green).
    NAVY = colors.HexColor("#15315C")
    GREEN = colors.HexColor("#01A982")
    GREY = colors.HexColor("#5B6675")
    RED = colors.HexColor("#B23A3A")
    YELLOW = colors.HexColor("#C9A227")
    BLUE = colors.HexColor("#1E50A0")
    INK = colors.HexColor("#1F2937")
    HAIR = colors.HexColor("#E3E8EF")
    BOXBG = colors.HexColor("#FBFCFE")

    MARGIN = 0.5 * cm                       # house standard: 0.5 cm all sides
    USABLE = A4[0] - 2 * MARGIN             # usable text width; everything is clamped to this

    def pcolor(pr: str):
        return {"critical": RED, "high": YELLOW, "medium": BLUE, "low": GREEN}.get(str(pr).lower(), GREY)

    def hx(c) -> str:
        return "#%02X%02X%02X" % (int(c.red * 255), int(c.green * 255), int(c.blue * 255))

    def t(x):
        return escape(str(x)).replace("\n", "<br/>")

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN, topMargin=MARGIN, bottomMargin=MARGIN,
        title="CEGP Remediation Action Plan", author=generated_by,
    )

    ss = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=ss["Title"], fontSize=17, textColor=colors.white, alignment=TA_LEFT, spaceAfter=0, leading=20)
    subtitle_s = ParagraphStyle("st", parent=ss["Normal"], fontSize=8.5, textColor=colors.HexColor("#EAF6F1"), alignment=TA_LEFT, leading=11)
    card_h = ParagraphStyle("ch", parent=ss["Heading3"], fontSize=11, textColor=NAVY, spaceBefore=0, spaceAfter=4, leading=14)
    lbl = ParagraphStyle("lb", parent=ss["Normal"], fontSize=8, textColor=GREEN, spaceBefore=4, spaceAfter=1, leading=10, fontName="Helvetica-Bold")
    body = ParagraphStyle("bd", parent=ss["Normal"], fontSize=9, textColor=INK, leading=13, spaceAfter=1, alignment=TA_JUSTIFY)
    steps = ParagraphStyle("sp", parent=body, alignment=TA_LEFT)
    meta = ParagraphStyle("mt", parent=ss["Normal"], fontSize=8, textColor=GREY, leading=11, alignment=TA_LEFT)
    th = ParagraphStyle("th", parent=ss["Normal"], fontSize=8, textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_LEFT, leading=10)
    tv = ParagraphStyle("tv", parent=ss["Normal"], fontSize=12, textColor=NAVY, fontName="Helvetica-Bold", alignment=TA_LEFT, leading=14)

    # ---- Page number bottom-right (only footer content) ----
    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(GREY)
        canvas.drawRightString(A4[0] - MARGIN, MARGIN * 0.55, f"Page {doc_.page}")
        canvas.restoreState()

    # ---- Title band (single-cell, full usable width) ----
    title_tbl = Table([[Paragraph("CEGP — Remediation Action Plan", title_s)],
                       [Paragraph(f"Cyber Exposure Governance Platform &nbsp;|&nbsp; {len(plan_df)} action item(s) "
                                  f"&nbsp;|&nbsp; generated {datetime.now():%Y-%m-%d %H:%M} by {t(generated_by)}", subtitle_s)]],
                      colWidths=[USABLE])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("LINEBELOW", (0, -1), (-1, -1), 2, GREEN),
        ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (0, 0), 10), ("BOTTOMPADDING", (0, -1), (-1, -1), 9),
    ]))

    # ---- Summary table (fixed width = USABLE, four equal columns) ----
    total = len(plan_df)
    crit_high = int(plan_df["Priority"].astype(str).isin(["Critical", "High"]).sum()) if total else 0
    kev = int((plan_df["KEV"].astype(str).str.lower() == "yes").sum()) if total else 0
    breached = int((plan_df["SLA Status"].astype(str) == "Breached").sum()) if total else 0
    cw = USABLE / 4.0
    summary_tbl = Table(
        [[Paragraph("ACTION ITEMS", th), Paragraph("CRITICAL / HIGH", th), Paragraph("CISA KEV (FIX FIRST)", th), Paragraph("SLA BREACHED", th)],
         [Paragraph(str(total), tv), Paragraph(str(crit_high), tv), Paragraph(str(kev), tv), Paragraph(str(breached), tv)]],
        colWidths=[cw, cw, cw, cw])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EEF1F5")),
        ("BOX", (0, 0), (-1, -1), 0.5, HAIR),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    story = [title_tbl, Spacer(1, 8), summary_tbl, Spacer(1, 10)]

    # ---- One fixed-width boxed card per exposure ----
    for _, r in plan_df.iterrows():
        pr = str(r.get("Priority", ""))
        pc = pcolor(pr)
        head = (f'<font color="{hx(pc)}"><b>[{t(pr)}]</b></font> &nbsp;'
                f'{t(r.get("Rank"))}. {t(r.get("CVE"))} on {t(r.get("Asset"))} '
                f'&nbsp;&middot;&nbsp; score {t(r.get("Exposure Score"))} '
                f'&nbsp;&middot;&nbsp; KEV {t(r.get("KEV"))} &nbsp;&middot;&nbsp; SLA {t(r.get("SLA Status"))}')
        meta_line = (f"Owner: {t(r.get('Owner') or '—')} &nbsp;|&nbsp; Due: {t(r.get('Due Date') or '—')} "
                     f"&nbsp;|&nbsp; Control area: {t(r.get('Control Area') or '—')}"
                     + (f" &nbsp;|&nbsp; Escalation: {t(r.get('Escalation'))}" if str(r.get("Escalation") or "").strip() else ""))
        inner = [
            Paragraph(head, card_h),
            Paragraph("WHAT THE VULNERABILITY IS", lbl), Paragraph(t(r.get("Vulnerability Explanation")), body),
            Paragraph("WHY IT MATTERS / POSSIBLE IMPACT", lbl), Paragraph(t(r.get("Possible Impact")), body),
            Paragraph("REMEDIATION STEPS", lbl), Paragraph(t(r.get("Remediation Steps")), steps),
            Paragraph("INTERIM MITIGATION", lbl), Paragraph(t(r.get("Interim Mitigation")), steps),
            Spacer(1, 3), Paragraph(meta_line, meta),
        ]
        # Single-cell table = the card box. colWidths=[USABLE] guarantees it never exceeds the page.
        card = Table([[inner]], colWidths=[USABLE])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), BOXBG),
            ("BOX", (0, 0), (-1, -1), 0.5, HAIR),
            ("LINEBEFORE", (0, 0), (0, -1), 3, pc),     # priority-coloured left stripe
            ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(KeepTogether([card, Spacer(1, 7)]))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


def to_executive_pdf_bytes(df: pd.DataFrame, generated_by: str = "system") -> bytes:
    """Export the Executive view (visualisations + data analysis) as a consulting-grade PDF.

    Follows the house document standard: 0.5 cm margins, justified text, page number
    bottom-right only, fixed-width tables/boxes that never overflow, formal palette.
    Charts are rebuilt here (self-contained) and embedded as images; if static image
    export is unavailable, the PDF is still produced with the KPIs and data tables.
    """
    from datetime import datetime
    from xml.sax.saxutils import escape
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                                    Image as RLImage, KeepTogether)

    C_NAVY = colors.HexColor("#15315C")
    C_GREEN = colors.HexColor("#01A982")
    C_GREY = colors.HexColor("#5B6675")
    C_HAIR = colors.HexColor("#E3E8EF")
    C_INK = colors.HexColor("#1F2937")

    MARGIN = 0.5 * cm
    USABLE = A4[0] - 2 * MARGIN

    def t(x):
        return escape(str(x)).replace("\n", "<br/>")

    # ---- Render the executive charts with reportlab's native graphics ----
    # Uses ONLY reportlab (already required to build this PDF), so the charts render
    # wherever the PDF does - no kaleido, no matplotlib, nothing extra to install.
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
    from reportlab.graphics.charts.piecharts import Pie

    def _RC(h):
        return colors.HexColor(h)

    PMAP = {"Critical": _RC(RED), "High": _RC(YELLOW), "Medium": _RC(BLUE), "Low": _RC("#2E7D5B")}
    COLW = (USABLE - 8) / 2.0          # width of one 2-up cell
    CHART_W = COLW - 6                  # drawing width (fits inside the cell)
    CH = 190                           # 2-up chart height
    AX = _RC("#5B6675"); TITLE = _RC("#15315C"); GRID = _RC("#E7ECF2")

    def _title(d, w, txt):
        d.add(String(w / 2.0, d.height - 11, txt, fontName="Helvetica-Bold",
                     fontSize=9, fillColor=TITLE, textAnchor="middle"))

    def _vbar_chart(values, cats, color, title, w=CHART_W, h=CH):
        d = Drawing(w, h); _title(d, w, title)
        bc = VerticalBarChart()
        bc.x, bc.y = 30, 26
        bc.width, bc.height = w - 44, h - 52
        bc.data = [list(values)]
        bc.categoryAxis.categoryNames = [str(c) for c in cats]
        bc.categoryAxis.labels.fontSize = 6.5
        bc.categoryAxis.labels.fillColor = AX
        bc.categoryAxis.labels.dy = -2
        bc.categoryAxis.strokeColor = GRID
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 6.5
        bc.valueAxis.labels.fillColor = AX
        bc.valueAxis.strokeColor = GRID
        bc.valueAxis.gridStrokeColor = GRID
        bc.bars[0].fillColor = color
        bc.barSpacing = 1
        d.add(bc); return d

    def _hbar_chart(values, cats, color, title, w=CHART_W, h=CH):
        pairs = sorted(zip(cats, values), key=lambda p: p[1])     # ascending -> largest on top
        cats = [(str(c)[:20]) for c, _ in pairs]; values = [v for _, v in pairs]
        d = Drawing(w, h); _title(d, w, title)
        bc = HorizontalBarChart()
        bc.x, bc.y = 100, 16
        bc.width, bc.height = w - 112, h - 38
        bc.data = [list(values)]
        bc.categoryAxis.categoryNames = cats
        bc.categoryAxis.labels.fontSize = 6
        bc.categoryAxis.labels.fillColor = AX
        bc.categoryAxis.labels.boxAnchor = "e"
        bc.categoryAxis.strokeColor = GRID
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 6.5
        bc.valueAxis.labels.fillColor = AX
        bc.valueAxis.strokeColor = GRID
        bc.valueAxis.gridStrokeColor = GRID
        bc.bars[0].fillColor = color
        bc.barSpacing = 1
        d.add(bc); return d

    def _pie_chart(values, labels, slice_colors, title, w=CHART_W, h=CH):
        d = Drawing(w, h); _title(d, w, title)
        pie = Pie()
        pie.width = pie.height = h - 58
        pie.x = (w - pie.width) / 2.0; pie.y = 16
        pie.data = list(values)
        pie.labels = ["%s %d" % (l, v) for l, v in zip(labels, values)]
        pie.sideLabels = 1
        pie.slices.strokeColor = colors.white
        pie.slices.strokeWidth = 1
        pie.slices.fontSize = 6.5
        pie.slices.fontColor = AX
        for k, c in enumerate(slice_colors):
            pie.slices[k].fillColor = c
        d.add(pie); return d

    def _matrix_drawing(counts, w=USABLE, h=230):
        d = Drawing(w, h)
        _title(d, w, "Executive Risk Matrix \u2014 Severity \u00d7 Exploitation Likelihood")
        left, bottom, top, right = 96, 36, 22, 8
        gw = w - left - right; gh = h - bottom - top
        cw = gw / 4.0; chh = gh / 4.0
        stops = [(0.0, (0xDC, 0xEF, 0xE6)), (0.34, (0xEF, 0xF3, 0xD9)), (0.5, (0xF6, 0xEC, 0xC9)),
                 (0.72, (0xEF, 0xC9, 0xA6)), (1.0, (0xEB, 0xB4, 0xB4))]

        def _col(v):
            for k in range(len(stops) - 1):
                v0, c0 = stops[k]; v1, c1 = stops[k + 1]
                if v <= v1:
                    f = (v - v0) / (v1 - v0) if v1 > v0 else 0
                    return colors.Color((c0[0] + (c1[0] - c0[0]) * f) / 255.0,
                                        (c0[1] + (c1[1] - c0[1]) * f) / 255.0,
                                        (c0[2] + (c1[2] - c0[2]) * f) / 255.0)
            return colors.Color(*[x / 255.0 for x in stops[-1][1]])

        for yi in range(4):
            for xi in range(4):
                x = left + xi * cw; y = bottom + yi * chh
                d.add(Rect(x, y, cw - 3, chh - 3, fillColor=_col((xi + yi) / 6.0),
                           strokeColor=colors.white, strokeWidth=1.5))
                if counts[yi][xi]:
                    d.add(String(x + (cw - 3) / 2.0, y + (chh - 3) / 2.0 - 4, str(counts[yi][xi]),
                                 fontName="Helvetica-Bold", fontSize=11, fillColor=TITLE, textAnchor="middle"))
        for yi, lab in enumerate(["Low <0.50", "Med 0.50-0.79", "High 0.80-0.94", "V.High \u22650.95"]):
            d.add(String(left - 6, bottom + yi * chh + chh / 2.0 - 3, lab, fontName="Helvetica",
                         fontSize=6.5, fillColor=AX, textAnchor="end"))
        for xi, lab in enumerate(["Low", "Medium", "High", "Critical"]):
            d.add(String(left + xi * cw + cw / 2.0, bottom - 13, lab, fontName="Helvetica",
                         fontSize=7, fillColor=AX, textAnchor="middle"))
        d.add(String(left + gw / 2.0, 5, "CVSS severity \u2192", fontName="Helvetica",
                     fontSize=7, fillColor=AX, textAnchor="middle"))
        return d

    draw, charts_ok, chart_err = {}, False, ""
    try:
        def _cb(v):
            v = float(v or 0)
            return 3 if v >= 9 else 2 if v >= 7 else 1 if v >= 4 else 0

        def _eb(p):
            p = float(p or 0)
            return 3 if p >= 0.95 else 2 if p >= 0.80 else 1 if p >= 0.50 else 0

        counts = [[0] * 4 for _ in range(4)]
        for _, r in df.iterrows():
            counts[_eb(r.get("epss_percentile", 0))][_cb(r.get("cvss_score", 0))] += 1
        draw["matrix"] = _matrix_drawing(counts)

        order = ["Critical", "High", "Medium", "Low"]
        vc = df["priority"].value_counts()
        keep = [(o, int(vc.get(o, 0))) for o in order if int(vc.get(o, 0)) > 0]
        draw["donut"] = _pie_chart([v for _, v in keep], [o for o, _ in keep],
                                   [PMAP[o] for o, _ in keep], "Priority Mix")

        tb = df.groupby("business_process", dropna=False)["final_score"].sum().sort_values(ascending=False).head(8)
        draw["bp"] = _hbar_chart(list(tb.values), list(tb.index), _RC(GREEN), "Top Business Processes by Risk")

        kv = df["kev_status"].value_counts()
        draw["kev"] = _vbar_chart([int(kv.get("Yes", 0)), int(kv.get("No", 0))],
                                  ["Known Exploited", "Not in KEV"], _RC(RED), "CISA KEV vs Non-KEV")

        ep = pd.to_numeric(df["epss_percentile"], errors="coerce").dropna()
        ebins = [0] * 10
        for v in ep:
            ebins[min(int(float(v) * 10), 9)] += 1
        draw["epss"] = _vbar_chart(ebins, ["0", "", ".2", "", ".4", "", ".6", "", ".8", ""],
                                   _RC(BLUE), "EPSS Percentile Distribution")

        ev = df.groupby("environment")["final_score"].sum().sort_values(ascending=False)
        draw["env"] = _vbar_chart(list(ev.values), list(ev.index), _RC(GREEN), "Aggregate Risk by Environment")

        ta = df.groupby("asset_name")["final_score"].sum().sort_values(ascending=False).head(8)
        draw["assets"] = _hbar_chart(list(ta.values), list(ta.index), _RC(RED), "Top Riskiest Assets")

        charts_ok = len(draw) > 0
    except Exception as _e:
        charts_ok = False
        chart_err = str(_e)

    # ---- Document ----
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=MARGIN, rightMargin=MARGIN,
                            topMargin=MARGIN, bottomMargin=MARGIN,
                            title="CEGP Executive Cyber Exposure Summary", author=generated_by)
    ss = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=ss["Title"], fontSize=17, textColor=colors.white, alignment=TA_LEFT, leading=20)
    sub_s = ParagraphStyle("s", parent=ss["Normal"], fontSize=8.5, textColor=colors.HexColor("#EAF6F1"), alignment=TA_LEFT, leading=11)
    sec = ParagraphStyle("sec", parent=ss["Heading3"], fontSize=11, textColor=C_NAVY, spaceBefore=8, spaceAfter=4)
    body = ParagraphStyle("bd", parent=ss["Normal"], fontSize=9, textColor=C_INK, leading=13, alignment=TA_JUSTIFY)
    cap = ParagraphStyle("cap", parent=ss["Normal"], fontSize=8, textColor=C_GREY, leading=10, spaceBefore=2)
    th = ParagraphStyle("th", parent=ss["Normal"], fontSize=8, textColor=colors.white, fontName="Helvetica-Bold", leading=10)
    tv = ParagraphStyle("tv", parent=ss["Normal"], fontSize=14, textColor=C_NAVY, fontName="Helvetica-Bold", leading=16)
    cell = ParagraphStyle("cell", parent=ss["Normal"], fontSize=8.5, textColor=C_INK, leading=11)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 8); canvas.setFillColor(C_GREY)
        canvas.drawRightString(A4[0] - MARGIN, MARGIN * 0.55, f"Page {doc_.page}")
        canvas.restoreState()

    # KPI metrics
    s = executive_summary_report(df)
    m = dict(zip(s["Metric"], s["Value"]))

    title_tbl = Table([[Paragraph("CEGP \u2014 Executive Cyber Exposure Summary", title_s)],
                       [Paragraph(f"Cyber Exposure Governance Platform &nbsp;|&nbsp; {len(df)} exposures &nbsp;|&nbsp; "
                                  f"generated {datetime.now():%Y-%m-%d %H:%M} by {t(generated_by)}", sub_s)]],
                      colWidths=[USABLE])
    title_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), C_NAVY), ("LINEBELOW", (0, -1), (-1, -1), 2, C_GREEN),
                                   ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                                   ("TOPPADDING", (0, 0), (0, 0), 10), ("BOTTOMPADDING", (0, -1), (-1, -1), 9)]))

    kpis = [("TOTAL", m.get("Total Exposures", len(df))), ("CRITICAL", m.get("Critical Exposures", 0)),
            ("HIGH", m.get("High Exposures", 0)), ("CISA KEV", m.get("CISA KEV Exposures", 0)),
            ("SLA BREACHED", m.get("SLA Breached Items", 0)), ("AVG SCORE", m.get("Average Risk Score", 0))]

    def _kfmt(label, v):
        if label == "AVG SCORE":
            try:
                return f"{float(v):.2f}"
            except Exception:
                return str(v)
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else str(v)
        except Exception:
            return str(v)

    cwk = USABLE / 6.0
    kpi_tbl = Table([[Paragraph(k, th) for k, _ in kpis], [Paragraph(_kfmt(k, v), tv) for k, v in kpis]],
                    colWidths=[cwk] * 6)
    kpi_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), C_NAVY), ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EEF1F5")),
                                 ("BOX", (0, 0), (-1, -1), 0.5, C_HAIR), ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
                                 ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                                 ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                                 ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))

    story = [title_tbl, Spacer(1, 8), kpi_tbl, Spacer(1, 6)]

    def _cell(name):
        d = draw.get(name)
        return d if d is not None else Paragraph("(chart unavailable)", cap)

    if charts_ok and draw:
        story += [Paragraph("Risk Matrix \u2014 fix-first is the top-right band", sec),
                  _cell("matrix"),
                  Paragraph("Each cell counts exposures in that CVSS severity \u00d7 EPSS likelihood band.", cap),
                  Spacer(1, 6), Paragraph("Posture Breakdown", sec)]
        colw = (USABLE - 8) / 2.0
        for a, b in [("donut", "bp"), ("kev", "epss"), ("env", "assets")]:
            row = Table([[_cell(a), _cell(b)]], colWidths=[colw, colw])
            row.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                                     ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                                     ("VALIGN", (0, 0), (-1, -1), "TOP")]))
            story.append(row)
    else:
        msg = "Charts could not be rendered; KPIs and data tables follow."
        if chart_err:
            msg += " (reason: %s)" % t(chart_err)
        story.append(Paragraph(msg, body))

    # ---- Data analysis: top exposures table (fixed width, wrapped) ----
    story += [Spacer(1, 6), Paragraph("Top 15 Business-Critical Exposures", sec)]
    cols = [("Asset", "asset_name", 0.22), ("Business Process", "business_process", 0.22),
            ("CVE", "cve_id", 0.16), ("Priority", "priority", 0.12), ("Score", "final_score", 0.10),
            ("KEV", "kev_status", 0.18)]
    head = [Paragraph(h, th) for h, _, _ in cols]
    rows = [head]
    top = df.sort_values("final_score", ascending=False).head(15)
    for _, r in top.iterrows():
        rows.append([Paragraph(t(r.get(c, "")), cell) for _, c, _ in cols])
    data_tbl = Table(rows, colWidths=[USABLE * w for _, _, w in cols], repeatRows=1)
    data_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
                                  ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFCFE")]),
                                  ("BOX", (0, 0), (-1, -1), 0.5, C_HAIR), ("INNERGRID", (0, 0), (-1, -1), 0.4, C_HAIR),
                                  ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                                  ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                                  ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(data_tbl)

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


def run_assessment(vuln_df, asset_df, ids_df, exceptions_df, import_type, use_live_feeds, nvd_api_key):
    vuln_df = normalize_vulnerability_input(vuln_df, import_type=import_type)
    vuln_df, vuln_errors, vuln_warnings = validate_vulnerability_df(vuln_df)
    asset_df, asset_errors, asset_warnings = validate_asset_df(asset_df)
    ids_df, ids_errors, ids_warnings = validate_ids_df(ids_df) if ids_df is not None else (pd.DataFrame(), [], [])

    errors = vuln_errors + asset_errors + ids_errors
    warnings = vuln_warnings + asset_warnings + ids_warnings
    if errors:
        return None, errors, warnings

    policy = load_risk_policy(DATA_DIR / "risk_policy.json")

    df = enrich_with_kev(vuln_df, use_live=use_live_feeds)
    df = enrich_with_epss(df, use_live=use_live_feeds)
    df = enrich_with_nvd(df, use_live=use_live_feeds, api_key=nvd_api_key or None)

    df = enrich_with_asset_context(df, asset_df)
    df = calculate_business_impact(df)
    df = add_network_exposure(df)
    df = correlate_ids_alerts(df, ids_df)
    df = add_privacy_impact(df)

    df = calculate_threat_intelligence_score(df, policy)
    df = calculate_final_score(df, policy, include_sla_score=False)
    df = assign_remediation_governance(df, policy)
    df = calculate_final_score(df, policy, include_sla_score=True)
    df = assign_remediation_governance(df, policy)

    df = add_remediation_playbooks(df)
    df = apply_exceptions(df, exceptions_df)
    df = add_control_mapping(df)
    df["score_drivers"] = df.apply(summarize_score_drivers, axis=1)

    priority_order = {"Critical": 1, "High": 2, "Medium": 3, "Low": 4}
    df["_priority_order"] = df["priority"].map(priority_order).fillna(5)
    df = df.sort_values(["_priority_order", "final_score"], ascending=[True, False]).drop(columns=["_priority_order"])

    log_event("Assessment completed", details=f"{len(df)} exposure rows assessed", operator=operator())
    try:
        save_risk_snapshot(df, DATA_DIR / "risk_snapshots.csv")
    except Exception:
        pass

    return df, errors, warnings


# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------
inject_theme()
render_header()

with st.sidebar:
    st.markdown(
        '<div class="cegp-sidebrand"><span style="font-size:1.5rem">\U0001F6E1\uFE0F</span>'
        '<div><div class="b1">CEGP Console</div><div class="b2">Cyber Exposure Governance</div></div></div>',
        unsafe_allow_html=True,
    )
    if st.button("\U0001F4C4  Project Documentation", use_container_width=True, key="docs_btn"):
        show_documentation()
    if st.button("\U0001F680  Future Scope  \nImplementation using GCP", use_container_width=True, key="gcp_btn"):
        show_gcp_guide()
    st.divider()

    st.markdown('<div class="cegp-sidelabel">Assessment Setup</div>', unsafe_allow_html=True)
    st.session_state.operator = st.text_input("Operator name", value=st.session_state.get("operator", "analyst"),
                                               help="Stamped against every action in the audit log.")
    use_samples = st.toggle("Use bundled sample files", value=True)
    if not use_samples:
        vuln_file = st.file_uploader("Vulnerability file (CSV / Excel)", type=["csv", "xlsx", "xls"])
        asset_file = st.file_uploader("Asset inventory (CSV / Excel)", type=["csv", "xlsx", "xls"])
        ids_file = st.file_uploader("IDS/IPS alerts (CSV / Excel)", type=["csv", "xlsx", "xls"])
        exception_file = st.file_uploader("Risk exceptions (CSV / Excel, optional)", type=["csv", "xlsx", "xls"])
    else:
        vuln_file = asset_file = ids_file = exception_file = None

    run_button = st.button("Run Cyber Exposure Assessment", type="primary", use_container_width=True)

    # ---- Bottom of sidebar: optional / advanced, shown only if expanded ----
    st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
    with st.expander("Advanced options", expanded=False):
        import_type = st.selectbox(
            "Vulnerability input type",
            ["Standard Template", "Scanner Export - Basic", "OpenVAS-like CSV", "Manual CVE List"],
            index=0,
            help="How to parse the uploaded file. Use Standard Template for the bundled data and matching files.",
        )
        use_live_feeds = st.toggle("Use live public feeds when available", value=False,
                                   help="When off, bundled offline threat-intel is used (recommended for a stable demo).")
        nvd_api_key = st.text_input("NVD API key (optional)", type="password",
                                    help="Only used with live feeds on; raises NVD rate limits.")
    render_sidebar_concepts()
    with st.expander("About this project", expanded=False):
        st.markdown(
            "**Cyber Exposure Governance Platform** \u2014 v1\n\n"
            "An M.Sc. Cyber Security project by **Dwaipayan Mojumder** and **Deblina Das**, "
            "under the guidance of **Prof. Sanjay Pal**.\n\n"
            "A risk-based decision-support tool: it correlates vulnerability intelligence with "
            "asset, network, IDS/IPS and privacy context to prioritise remediation and produce "
            "audit-ready, tamper-evident reports."
        )


if "assessment_df" not in st.session_state:
    st.session_state.assessment_df = None
if "last_hashes" not in st.session_state:
    st.session_state.last_hashes = {}
if "dq_notices" not in st.session_state:
    st.session_state.dq_notices = []


if run_button:
    try:
        vuln_df = load_csv(vuln_file, DATA_DIR / "sample_input.csv")
        asset_df = load_csv(asset_file, DATA_DIR / "asset_inventory.csv")
        ids_df = load_csv(ids_file, DATA_DIR / "ids_alerts.csv")
        exceptions_df = load_csv(exception_file, DATA_DIR / "risk_exceptions.csv") if (exception_file is not None or (DATA_DIR / "risk_exceptions.csv").exists()) else load_exceptions(DATA_DIR / "risk_exceptions.csv")

        with st.spinner("Running cyber exposure assessment..."):
            result_df, errors, warnings = run_assessment(
                vuln_df, asset_df, ids_df, exceptions_df,
                import_type=import_type, use_live_feeds=use_live_feeds, nvd_api_key=nvd_api_key,
            )

        # Data-quality warnings are not printed at the top of the page; they are
        # stored and surfaced under the Audit Log tab so the first screen stays clean.
        st.session_state.dq_notices = warnings

        if errors:
            for error in errors:
                st.error(error)
        else:
            st.session_state.assessment_df = result_df
            # Archive this run to history (full results + input fingerprints). Never block the run.
            try:
                run_id = save_run(
                    result_df, operator=operator(), data_dir=str(DATA_DIR),
                    inputs={
                        "vuln_rows": len(vuln_df),
                        "asset_rows": len(asset_df),
                        "ids_rows": (len(ids_df) if ids_df is not None else 0),
                        "vuln_fingerprint": fingerprint_df(vuln_df),
                        "asset_fingerprint": fingerprint_df(asset_df),
                        "ids_fingerprint": fingerprint_df(ids_df) if ids_df is not None else "",
                    },
                )
                log_event("Assessment archived to history", details=f"run_id={run_id}", operator=operator())
            except Exception:
                pass
            note = f" ({len(warnings)} data-quality notice{'s' if len(warnings) != 1 else ''} \u2014 see Audit Log tab)" if warnings else ""
            st.success("Assessment completed successfully." + note)
    except Exception as exc:
        st.error(f"Assessment failed: {exc}")
        log_event("Assessment failed", details=str(exc), operator=operator())


df = st.session_state.assessment_df

if df is None:
    section("Welcome", GREEN)
    st.markdown(
        "This platform turns raw vulnerability data into **business-aligned remediation decisions**. "
        "It enriches CVEs with threat intelligence (CISA KEV, EPSS, CVSS) and correlates asset, network, "
        "IDS/IPS and privacy context to produce an explainable exposure score, SLA governance, and "
        "tamper-evident reports."
    )
    c1, c2, c3 = st.columns(3)
    kpi_card(c1, "Prioritise", "Explainable 0\u2013100 score", GREEN, GREEN_D, "Threat + business + network + IDS + privacy")
    kpi_card(c2, "Govern", "SLA \u00b7 Owners \u00b7 Exceptions", BLUE, NAVY, "Due dates, escalation, risk acceptance")
    kpi_card(c3, "Evidence", "Audit log + HMAC", YELLOW, "#8A6D0E", "Tamper-evident, injection-safe reports")
    st.info("Set your operator name in the sidebar, then click **Run Cyber Exposure Assessment** to begin.")
    st.stop()


# ---- Executive KPI cards ----
summary = executive_summary_report(df)
m = dict(zip(summary["Metric"], summary["Value"]))

section("Cyber Exposure Posture", NAVY)
r1 = st.columns(4)
kpi_card(r1[0], "Total Exposures", m["Total Exposures"], GREEN, NAVY)
kpi_card(r1[1], "Critical", m["Critical Exposures"], RED, RED)
kpi_card(r1[2], "CISA KEV", m["CISA KEV Exposures"], YELLOW, "#8A6D0E")
kpi_card(r1[3], "SLA Breached", m["SLA Breached Items"], RED, RED)

r2 = st.columns(4)
kpi_card(r2[0], "IDS-Correlated", m["IDS-Correlated Exposures"], GREEN, GREEN_D)
kpi_card(r2[1], "Privacy Critical/High", m["Privacy Critical/High Exposures"], GREY, NAVY)
kpi_card(r2[2], "Average Risk Score", m["Average Risk Score"], GREEN, GREEN_D)
kpi_card(r2[3], "High", m["High Exposures"], YELLOW, "#8A6D0E")

st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)

tabs = st.tabs(
    [
        "Executive View", "Analyst View", "Network Exposure", "IDS/IPS Correlation",
        "Privacy Impact", "Remediation Governance", "Action Plan", "Simulation", "Reports & Integrity", "Audit Log",
        "Assessment History",
    ]
)

with tabs[0]:
    section("Executive Cyber Exposure View", NAVY)

    # Hero: executive risk matrix — exposures binned by severity x exploitation likelihood.
    pc = df["priority"].value_counts().reset_index()
    pc.columns = ["Priority", "Count"]

    cvss_labels = ["Low<br>0\u20133.9", "Medium<br>4\u20136.9", "High<br>7\u20138.9", "Critical<br>9\u201310"]
    epss_labels = ["Low<br>&lt;0.50", "Medium<br>0.50\u20130.79", "High<br>0.80\u20130.94", "Very High<br>\u22650.95"]

    def _cvss_band(v):
        v = float(v or 0)
        return 3 if v >= 9 else 2 if v >= 7 else 1 if v >= 4 else 0

    def _epss_band(p):
        p = float(p or 0)
        return 3 if p >= 0.95 else 2 if p >= 0.80 else 1 if p >= 0.50 else 0

    counts = [[0] * 4 for _ in range(4)]   # [epss_band][cvss_band]
    for _, r in df.iterrows():
        counts[_epss_band(r.get("epss_percentile", 0))][_cvss_band(r.get("cvss_score", 0))] += 1

    tier = [[(x + y) / 6 for x in range(4)] for y in range(4)]            # 0 (cool) .. 1 (hot)
    text = [[str(counts[y][x]) if counts[y][x] else "" for x in range(4)] for y in range(4)]

    matrix = go.Figure(go.Heatmap(
        z=tier, x=cvss_labels, y=epss_labels,
        text=text, texttemplate="%{text}", textfont=dict(size=20, color="#15315C", family="Segoe UI"),
        customdata=counts,
        hovertemplate="Severity: %{x}<br>Likelihood: %{y}<br>Exposures: %{customdata}<extra></extra>",
        xgap=5, ygap=5, showscale=False,
        colorscale=[[0.0, "#DCEFE6"], [0.34, "#EFF3D9"], [0.5, "#F6ECC9"], [0.72, "#EFC9A6"], [1.0, "#EBB4B4"]],
        zmin=0, zmax=1,
    ))
    matrix = style_fig(matrix)
    matrix.update_layout(
        title=dict(text="Executive Risk Matrix \u2014 Severity \u00d7 Exploitation Likelihood", font=dict(size=16, color=NAVY)),
        height=430, margin=dict(t=56, l=10, r=10, b=10),
    )
    matrix.update_xaxes(title="CVSS severity \u2192", side="bottom", showgrid=False, tickfont=dict(size=12, color=GREY))
    matrix.update_yaxes(title="EPSS likelihood \u2192", showgrid=False, tickfont=dict(size=12, color=GREY))
    st.plotly_chart(matrix, use_container_width=True)
    st.caption("Each cell shows the number of exposures in that severity \u00d7 likelihood band. The **top-right** band (high severity **and** high exploitation likelihood) is the fix-first zone; the bottom-left is monitor-and-defer.")


    r1l, r1r = st.columns(2)
    with r1l:
        donut = px.pie(pc, names="Priority", values="Count", hole=0.55, title="Priority Mix",
                       color="Priority", color_discrete_map=PRIORITY_COLORS,
                       category_orders={"Priority": ["Critical", "High", "Medium", "Low"]})
        donut.update_traces(textinfo="label+percent", textfont_size=12)
        donut = style_fig(donut)
        donut.update_layout(showlegend=False)
        st.plotly_chart(donut, use_container_width=True)
    with r1r:
        tb = df.groupby("business_process", dropna=False)["final_score"].sum().sort_values(ascending=False).head(10).reset_index()
        tbfig = px.bar(tb, x="final_score", y="business_process", orientation="h", title="Top Business Processes by Aggregate Risk")
        tbfig = style_fig(tbfig, GREEN)
        tbfig.update_yaxes(autorange="reversed", title="")
        tbfig.update_xaxes(title="Aggregate risk score")
        st.plotly_chart(tbfig, use_container_width=True)

    r2l, r2r = st.columns(2)
    with r2l:
        kc = df["kev_status"].value_counts().reindex(["Yes", "No"]).fillna(0).reset_index()
        kc.columns = ["KEV", "Count"]
        kc["KEV"] = kc["KEV"].map({"Yes": "Known Exploited (KEV)", "No": "Not in KEV"})
        kfig = px.bar(kc, x="KEV", y="Count", color="KEV", title="CISA KEV vs Non-KEV Exposures",
                      color_discrete_map={"Known Exploited (KEV)": RED, "Not in KEV": GREY})
        kfig = style_fig(kfig)
        kfig.update_layout(showlegend=False)
        st.plotly_chart(kfig, use_container_width=True)
    with r2r:
        hfig = px.histogram(df, x="epss_percentile", nbins=20, title="EPSS Percentile Distribution")
        hfig = style_fig(hfig, BLUE)
        hfig.update_layout(bargap=0.05)
        hfig.update_xaxes(title="EPSS percentile (exploitation likelihood)")
        hfig.update_yaxes(title="Exposures")
        st.plotly_chart(hfig, use_container_width=True)

    r3l, r3r = st.columns(2)
    with r3l:
        ev = df.groupby("environment")["final_score"].sum().sort_values(ascending=False).reset_index()
        efig = px.bar(ev, x="environment", y="final_score", title="Aggregate Risk by Environment")
        efig = style_fig(efig, GREEN)
        efig.update_yaxes(title="Total risk score")
        st.plotly_chart(efig, use_container_width=True)
    with r3r:
        ta = df.groupby("asset_name")["final_score"].sum().sort_values(ascending=False).head(10).reset_index()
        tafig = px.bar(ta, x="final_score", y="asset_name", orientation="h", title="Top 10 Riskiest Assets")
        tafig = style_fig(tafig, RED)
        tafig.update_yaxes(autorange="reversed", title="")
        tafig.update_xaxes(title="Aggregate risk score")
        st.plotly_chart(tafig, use_container_width=True)

    section("Top 10 Business-Critical Exposures", GREEN)
    cols = ["asset_name", "cve_id", "priority", "final_score"]
    render_table(df[cols].head(10))

    section("Download Executive Summary", NAVY)
    st.caption("A consulting-format PDF of this Executive view \u2014 risk matrix, posture charts, KPIs, and the "
               "top business-critical exposures \u2014 signed for integrity.")
    try:
        exec_pdf = to_executive_pdf_bytes(df, generated_by=operator())
        st.download_button("\u2b07\ufe0f  Download Executive Summary (PDF)", data=exec_pdf,
                           file_name="cegp_executive_summary.pdf", mime="application/pdf")
        st.caption("HMAC-SHA256")
        st.code(generate_sha256(exec_pdf), language="text")
    except ImportError:
        st.warning("PDF export needs the `reportlab` package (and `matplotlib` for embedded charts). "
                   "Add them to requirements.txt and rebuild.")
    except Exception as exc:
        st.warning(f"Executive PDF unavailable: {exc}")

with tabs[1]:
    section("Security Analyst Exposure Queue", NAVY)
    st.caption("Threat & scoring lens \u2014 severity, exploitation likelihood, and why each exposure scored as it did. "
               "Network, IDS, privacy, and remediation detail live in their own tabs.")
    analyst_cols = [
        "asset_name", "application_name", "cve_id", "severity", "cvss_score",
        "epss_score", "epss_percentile", "kev_status", "score_drivers", "priority", "final_score",
    ]
    sub = df[[c for c in analyst_cols if c in df.columns]]
    render_table(sub)

with tabs[2]:
    section("Network Exposure", NAVY)
    c = st.columns(2)
    with c[0]:
        zs = df.groupby("network_zone")["final_score"].sum().sort_values(ascending=False).reset_index()
        st.plotly_chart(style_fig(px.bar(zs, x="network_zone", y="final_score", title="Risk by Network Zone"), GREEN), use_container_width=True)
    with c[1]:
        if "asset_type" in df.columns:
            at = df.groupby("asset_type")["final_score"].sum().sort_values(ascending=False).reset_index()
            st.plotly_chart(style_fig(px.bar(at, x="asset_type", y="final_score", title="Risk by Asset Type"), BLUE), use_container_width=True)
    cols = ["asset_name", "network_zone", "open_ports", "firewall_status", "vpn_required",
            "network_exposure_level", "network_exposure_reason", "final_score"]
    render_table(df[cols].sort_values("final_score", ascending=False))

with tabs[3]:
    section("IDS/IPS Correlation", NAVY)
    c = st.columns(2)
    with c[0]:
        isum = df.groupby("highest_alert_severity")["asset_id"].count().reset_index()
        isum.columns = ["Highest Alert Severity", "Exposure Count"]
        st.plotly_chart(style_fig(px.bar(isum, x="Highest Alert Severity", y="Exposure Count", title="Correlated IDS/IPS Alert Severity"), YELLOW), use_container_width=True)
    with c[1]:
        cols = ["asset_name", "cve_id", "ids_alert_count", "highest_alert_severity",
                "exploit_attempt_detected", "signatures", "ids_correlation_reason"]
        ids_view = df.sort_values(["ids_alert_count", "final_score"], ascending=False)[cols]
        render_table(ids_view)

with tabs[4]:
    section("Privacy Impact", NAVY)
    c = st.columns(2)
    with c[0]:
        ps = df.groupby("privacy_impact_level")["asset_id"].count().reset_index()
        ps.columns = ["Privacy Impact", "Count"]
        st.plotly_chart(style_fig(px.bar(ps, x="Privacy Impact", y="Count", title="Privacy Impact Distribution"), GREEN), use_container_width=True)
    with c[1]:
        cols = ["asset_name", "data_type", "pii_present", "data_sensitivity", "encryption_status",
                "regulatory_impact", "privacy_impact_level", "privacy_reason"]
        priv_view = df.sort_values("privacy_impact_score", ascending=False)[cols] if "privacy_impact_score" in df.columns else df[cols]
        render_table(priv_view)

with tabs[5]:
    section("Remediation Governance", NAVY)
    c = st.columns(2)
    with c[0]:
        ss = df.groupby("sla_status")["asset_id"].count().reset_index()
        ss.columns = ["SLA Status", "Count"]
        sla_map = {"Breached": RED, "Due Soon": YELLOW, "Within SLA": GREEN}
        sla_fig = style_fig(px.bar(ss, x="SLA Status", y="Count", color="SLA Status", color_discrete_map=sla_map, title="SLA Governance Status"))
        sla_fig.update_layout(showlegend=False)
        st.plotly_chart(sla_fig, use_container_width=True)
    with c[1]:
        os_ = df.groupby("asset_owner")["final_score"].sum().sort_values(ascending=False).head(10).reset_index()
        ofig = px.bar(os_, x="final_score", y="asset_owner", orientation="h", title="Owner-wise Aggregate Risk")
        ofig = style_fig(ofig, GREEN)
        ofig.update_yaxes(autorange="reversed", title="")
        ofig.update_xaxes(title="Aggregate risk score")
        st.plotly_chart(ofig, use_container_width=True)

    if "control_area" in df.columns:
        ca = df["control_area"].dropna().str.split(",").explode().str.strip()
        ca = ca[ca != ""].value_counts().head(10).reset_index()
        ca.columns = ["Control Area", "Exposures"]
        cafig = px.bar(ca, x="Exposures", y="Control Area", orientation="h",
                       title="Exposures by Cybersecurity Control Area")
        cafig = style_fig(cafig, BLUE)
        cafig.update_yaxes(autorange="reversed", title="")
        st.plotly_chart(cafig, use_container_width=True)

    gov_cols = ["asset_name", "cve_id", "priority", "final_score", "asset_owner", "remediation_due_date",
                "days_remaining", "sla_status", "escalation_required", "escalation_reason",
                "exception_status", "exception_validity", "acceptance_reason", "accepted_by"]
    render_table(df[[c for c in gov_cols if c in df.columns]])

with tabs[6]:
    section("Remediation Action Plan", NAVY)
    st.caption(
        "Plain-language, prioritised **action items** for every identified exposure \u2014 what the "
        "vulnerability is, why it matters, exactly how to fix it, and what to do in the meantime. "
        "Worst-first ordering (KEV / Critical at the top)."
    )

    # Worst-first ordering: KEV, then priority, then exposure score.
    ap = df.copy()
    _po = {"Critical": 1, "High": 2, "Medium": 3, "Low": 4}
    ap["_kev"] = (ap.get("kev_status", "").astype(str).str.lower() == "yes").map({True: 0, False: 1})
    ap["_pr"] = ap.get("priority", "").map(_po).fillna(5)
    ap["_sc"] = pd.to_numeric(ap.get("final_score", 0), errors="coerce").fillna(0)
    ap = ap.sort_values(["_kev", "_pr", "_sc"], ascending=[True, True, False]).drop(columns=["_kev", "_pr", "_sc"])

    # ---- Action summary KPIs ----
    a1 = st.columns(4)
    kpi_card(a1[0], "Action Items", len(ap), GREEN, NAVY)
    kpi_card(a1[1], "Critical / High", int(ap["priority"].isin(["Critical", "High"]).sum()), RED, RED)
    kpi_card(a1[2], "CISA KEV (fix first)", int((ap.get("kev_status", "").astype(str).str.lower() == "yes").sum()), YELLOW, "#8A6D0E")
    kpi_card(a1[3], "SLA Breached", int((ap.get("sla_status", "") == "Breached").sum()), RED, RED)

    # ---- Filters ----
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    f = st.columns([2, 1, 1, 2])
    pr_opts = [p for p in ["Critical", "High", "Medium", "Low"] if p in set(ap["priority"].astype(str))]
    with f[0]:
        pr_sel = st.multiselect("Priority", pr_opts, default=pr_opts, key="ap_pr")
    with f[1]:
        kev_only = st.toggle("KEV only", value=False, key="ap_kev")
    with f[2]:
        breached_only = st.toggle("SLA breached only", value=False, key="ap_breach")
    with f[3]:
        query = st.text_input("Search asset / CVE / owner", key="ap_q", placeholder="e.g. CVE-2024 or web-server")

    view = ap[ap["priority"].astype(str).isin(pr_sel)] if pr_sel else ap.iloc[0:0]
    if kev_only:
        view = view[view.get("kev_status", "").astype(str).str.lower() == "yes"]
    if breached_only:
        view = view[view.get("sla_status", "") == "Breached"]
    if query.strip():
        q = query.strip().lower()

        def _s(frame, name):
            return frame[name].astype(str) if name in frame.columns else pd.Series([""] * len(frame), index=frame.index)

        hay = (_s(view, "asset_name") + " " + _s(view, "cve_id") + " "
               + _s(view, "asset_owner") + " " + _s(view, "application_name")).str.lower()
        view = view[hay.str.contains(q, na=False)]

    plan = compose_action_plan(view)

    # ---- Downloads (CSV / Excel / PDF), reflecting the current filter, HMAC-signed ----
    section("Download Action Plan", GREEN)
    st.caption(f"{len(plan)} action item(s) in the current view. Files are sanitised against CSV/formula injection and HMAC-signed (tamper-evident).")
    if plan.empty:
        st.info("No exposures match the current filters.")
    else:
        d1, d2, d3 = st.columns(3)
        csv_bytes = to_csv_bytes(plan)
        with d1:
            st.download_button("\u2b07\ufe0f  Download CSV", data=csv_bytes, file_name="cegp_action_plan.csv",
                               mime="text/csv", use_container_width=True)
            st.caption("HMAC-SHA256")
            st.code(generate_sha256(csv_bytes), language="text")
        with d2:
            try:
                xlsx_bytes = to_xlsx_bytes(plan)
                st.download_button("\u2b07\ufe0f  Download Excel (.xlsx)", data=xlsx_bytes, file_name="cegp_action_plan.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                st.caption("HMAC-SHA256")
                st.code(generate_sha256(xlsx_bytes), language="text")
            except Exception as exc:
                st.warning(f"Excel export unavailable: {exc}")
        with d3:
            try:
                pdf_bytes = to_pdf_bytes(plan, generated_by=operator())
                st.download_button("\u2b07\ufe0f  Download PDF", data=pdf_bytes, file_name="cegp_action_plan.pdf",
                                   mime="application/pdf", use_container_width=True)
                st.caption("HMAC-SHA256")
                st.code(generate_sha256(pdf_bytes), language="text")
            except ImportError:
                st.warning("PDF export needs the `reportlab` package. Add `reportlab>=4.0,<5.0` to requirements.txt and rebuild.")
            except Exception as exc:
                st.warning(f"PDF export unavailable: {exc}")
        log_event("Action plan exported", details=f"{len(plan)} items (CSV/XLSX/PDF prepared)", operator=operator())

    # ---- Expandable action cards ----
    section("Action Items", GREEN)
    if not plan.empty:
        max_cards = st.slider("Cards to display", 5, max(5, len(plan)), min(25, len(plan)), step=5, key="ap_n",
                              help="Limits on-screen cards for readability. Downloads above always include the full filtered set.")
        for _, c in plan.head(max_cards).iterrows():
            pr = str(c["Priority"])
            chip = PRIORITY_COLORS.get(pr, GREY)
            kev_tag = "  \u00b7  KEV" if str(c["KEV"]).lower() == "yes" else ""
            label = f"{c['Rank']}.  [{pr}]  {c['CVE']}  on  {c['Asset']}   \u00b7  score {c['Exposure Score']}  \u00b7  SLA {c['SLA Status']}{kev_tag}"
            with st.expander(label, expanded=False):
                st.markdown(
                    f"<div style='border-left:5px solid {chip};padding:2px 0 2px 12px;margin-bottom:6px'>"
                    f"<span style='color:{chip};font-weight:700'>{pr} priority</span>"
                    f" &nbsp;\u2014&nbsp; <span style='color:{GREY}'>Owner: {c['Owner'] or '\u2014'} &nbsp;|&nbsp; "
                    f"Due: {c['Due Date'] or '\u2014'} &nbsp;|&nbsp; Control area: {c['Control Area'] or '\u2014'}</span></div>",
                    unsafe_allow_html=True,
                )
                st.markdown(f"<span style='color:{GREEN_D};font-weight:700'>\U0001F50D What the vulnerability is</span>", unsafe_allow_html=True)
                st.markdown(c["Vulnerability Explanation"])
                st.markdown(f"<span style='color:{RED};font-weight:700'>\u26A0\uFE0F Why it matters / possible impact</span>", unsafe_allow_html=True)
                st.markdown(c["Possible Impact"])
                st.markdown(f"<span style='color:{NAVY};font-weight:700'>\U0001F6E0\uFE0F Remediation steps</span>", unsafe_allow_html=True)
                st.markdown(c["Remediation Steps"])
                st.markdown(f"<span style='color:{YELLOW};font-weight:700'>\U0001F9EF Interim mitigation (if you can't fix immediately)</span>", unsafe_allow_html=True)
                st.markdown(c["Interim Mitigation"])
                if str(c["Escalation"]).strip():
                    st.markdown(f"<span style='color:{GREY}'><b>Escalation:</b> {c['Escalation']}</span>", unsafe_allow_html=True)
        if len(plan) > max_cards:
            st.caption(f"Showing {max_cards} of {len(plan)} items. Increase the slider, refine filters, or use the downloads above for the full set.")

with tabs[7]:
    section("What-If Risk Reduction Simulator", NAVY)
    scenario = st.selectbox(
        "Select simulation scenario",
        [
            "Patch Top 5 Highest-Risk Exposures",
            "Patch All CISA KEV Exposures",
            "Isolate Internet-Facing Critical Assets",
            "Fix IDS-Correlated Exposures",
            "Encrypt Sensitive Unencrypted/Unknown Assets",
        ],
    )
    sim_summary, sim_df = run_simulation(df, scenario, load_risk_policy(DATA_DIR / "risk_policy.json"))
    log_event("Simulation executed", details=scenario, operator=operator())

    before_total = sim_summary.loc[sim_summary["Metric"] == "Total Risk", "Before"].iloc[0]
    after_total = sim_summary.loc[sim_summary["Metric"] == "Total Risk", "After"].iloc[0]
    try:
        reduction = round(float(before_total) - float(after_total), 2)
        pct = round((reduction / float(before_total)) * 100, 1) if float(before_total) else 0
    except Exception:
        reduction, pct = "-", "-"
    k = st.columns(3)
    kpi_card(k[0], "Total Risk Before", before_total, GREY, NAVY)
    kpi_card(k[1], "Total Risk After", after_total, GREEN, GREEN_D)
    kpi_card(k[2], "Risk Reduced", reduction, GREEN, GREEN_D, f"{pct}% lower" if pct != "-" else "")

    render_table(sim_summary, height=320)
    section("Simulated Exposure Queue", GREEN)
    ds = sim_df[["asset_name", "cve_id", "priority", "final_score", "simulated_score"]].sort_values("simulated_score", ascending=False)
    render_table(ds)

with tabs[8]:
    section("Reports, Ticket Export, and Integrity Verification", NAVY)
    st.caption("Reports are sanitised against CSV/formula injection and signed with HMAC-SHA256 (tamper-evident).")

    detailed_cols = [
        "asset_id", "asset_name", "application_name", "business_process", "cve_id", "severity", "cvss_score",
        "epss_score", "epss_percentile", "kev_status", "network_zone", "open_ports",
        "ids_alert_count", "privacy_impact_level", "final_score", "priority", "asset_owner",
        "remediation_due_date", "sla_status", "primary_action", "temporary_mitigation",
        "compensating_control", "validation_step", "control_area", "exception_status", "exception_validity",
    ]
    detailed_report = df[[c for c in detailed_cols if c in df.columns]].copy()
    executive_report = executive_summary_report(df)
    ticket_report = build_ticket_export(df)

    rc = st.columns(3)
    reports = {
        "detailed_remediation_report.csv": detailed_report,
        "executive_summary_report.csv": executive_report,
        "ticket_import_report.csv": ticket_report,
    }
    for idx, (filename, report_df) in enumerate(reports.items()):
        csv_bytes = to_csv_bytes(report_df)
        sig = generate_sha256(csv_bytes)
        st.session_state.last_hashes[filename] = sig
        with rc[idx]:
            st.markdown(f"**{filename.replace('_', ' ').replace('.csv', '').title()}**")
            st.download_button(f"Download CSV", data=csv_bytes, file_name=filename, mime="text/csv", use_container_width=True)
            st.caption("HMAC-SHA256 signature")
            st.code(sig, language="text")
            st.download_button("Download signature", data=f"{filename}\nHMAC-SHA256: {sig}\n".encode("utf-8"),
                               file_name=f"{filename}.hmac.txt", mime="text/plain", use_container_width=True)

    log_event("Reports prepared", details="Detailed, executive, and ticket export reports generated", operator=operator())

    section("Verify Report Integrity", GREEN)
    verify_file = st.file_uploader("Upload a report CSV to verify", type=["csv"], key="verify_report")
    original_hash = st.text_input("Paste original HMAC-SHA256 signature")
    if st.button("Verify Signature"):
        if verify_file is None or not original_hash.strip():
            st.warning("Upload a file and paste the original HMAC-SHA256 signature.")
        else:
            ok = verify_sha256(verify_file.getvalue(), original_hash)
            if ok:
                st.success("Verified: report integrity signature matches.")
                log_event("Report signature verified", details="Signature matched", operator=operator())
            else:
                st.error("Tampered or mismatch: signature does not match.")
                log_event("Report signature verification failed", details="Signature mismatch", operator=operator())

with tabs[9]:
    section("Audit Log and Risk Trend", NAVY)

    section("Data Quality Notices", YELLOW)
    notices = st.session_state.get("dq_notices", [])
    if notices:
        st.caption(f"{len(notices)} notice(s) raised during validation of the last assessment. These are non-blocking \u2014 rows are retained for review.")
        for n in notices:
            st.warning(n)
    else:
        st.caption("No data-quality notices from the last assessment.")

    audit_df = read_audit_log(DATA_DIR / "audit_log.csv")
    section("Audit Log", GREEN)
    render_table(audit_df.tail(100))

    section("Risk Trend Snapshots", GREEN)
    snapshots = read_risk_snapshots(DATA_DIR / "risk_snapshots.csv")
    if snapshots.empty:
        st.info("No risk snapshots found yet. Run assessment to create snapshots.")
    else:
        render_table(snapshots.tail(20), height=320)
        if len(snapshots) > 1:
            snap = snapshots.rename(columns={
                "critical_count": "Critical exposures",
                "high_count": "High exposures",
                "average_score": "Average score",
            })
            fig = px.line(
                snap, x="timestamp", y=["Critical exposures", "High exposures", "Average score"],
                title="Risk Trend Over Time", color_discrete_sequence=[RED, YELLOW, GREEN], markers=True,
            )
            fig = style_fig(fig)
            fig.update_layout(legend_title_text="")
            fig.update_xaxes(title="Snapshot time")
            fig.update_yaxes(title="Count / average score")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.caption("Run more than one assessment to see the risk trend build up over time.")

with tabs[10]:
    section("Assessment History", NAVY)
    _mode = storage_mode(str(DATA_DIR))
    _badge = ("Cloud Storage bucket" if _mode == "gcs" else "local disk")
    st.caption(
        "Every assessment run is archived here with its metrics, input fingerprints, and full "
        "results. Reopen a past run, compare it with the previous one, regenerate its Action Plan, "
        f"and re-download its signed reports \u2014 a reproducible, audit-defensible trail. "
        f"**Currently persisting to: {_badge}.**"
    )

    runs = list_runs(str(DATA_DIR))
    if runs.empty:
        st.info("No archived runs yet. Run a Cyber Exposure Assessment to create the first history entry.")
    else:
        reg_cols = ["run_id", "timestamp", "operator", "total_exposures", "critical", "high",
                    "medium", "low", "kev_count", "sla_breached", "average_score",
                    "vuln_rows", "asset_rows", "ids_rows"]
        render_table(runs[[c for c in reg_cols if c in runs.columns]], height=280)

        # Trend across runs
        if len(runs) > 1:
            tr = runs.sort_values("run_id").copy()
            for c in ["critical", "high", "average_score"]:
                tr[c] = pd.to_numeric(tr[c], errors="coerce")
            tr = tr.rename(columns={"critical": "Critical", "high": "High", "average_score": "Average score"})
            trfig = px.line(tr, x="timestamp", y=["Critical", "High", "Average score"], markers=True,
                            title="History \u2014 Critical / High / Average Score Across Runs",
                            color_discrete_sequence=[RED, YELLOW, GREEN])
            trfig = style_fig(trfig)
            trfig.update_layout(legend_title_text="")
            trfig.update_xaxes(title="Run")
            trfig.update_yaxes(title="Count / score")
            st.plotly_chart(trfig, use_container_width=True)

        section("Open a Past Run", GREEN)
        labels = [f"{r.run_id}  \u00b7  {r.timestamp}  \u00b7  {r.operator}  \u00b7  {r.total_exposures} exposures"
                  for r in runs.itertuples()]
        idx = st.selectbox("Select an archived run", range(len(runs)),
                           format_func=lambda i: labels[i], key="hist_pick")
        sel = runs.iloc[idx]
        run_df = load_run(str(sel["run_id"]), str(DATA_DIR))

        if run_df is None or run_df.empty:
            st.warning("The archived data file for this run could not be loaded.")
        else:
            h = st.columns(4)
            kpi_card(h[0], "Exposures", len(run_df), GREEN, NAVY)
            kpi_card(h[1], "Critical / High", int(run_df["priority"].astype(str).isin(["Critical", "High"]).sum()), RED, RED)
            kpi_card(h[2], "CISA KEV", int((run_df.get("kev_status", "").astype(str).str.lower() == "yes").sum()), YELLOW, "#8A6D0E")
            kpi_card(h[3], "SLA Breached", int((run_df.get("sla_status", "").astype(str) == "Breached").sum()), RED, RED)

            dash = "\u2014"
            st.caption(
                f"Input fingerprints \u2014 vuln: `{sel.get('vuln_fingerprint', '') or dash}` "
                f"\u00b7 asset: `{sel.get('asset_fingerprint', '') or dash}` "
                f"\u00b7 ids: `{sel.get('ids_fingerprint', '') or dash}`  "
                f"(these prove exactly which input data produced this run)."
            )

            # Compare with the previous run
            with st.expander("Compare with the previous run (new / resolved / persisting exposures)", expanded=False):
                order_ids = list(runs.sort_values("run_id")["run_id"].astype(str))
                cur_i = order_ids.index(str(sel["run_id"]))
                if cur_i == 0:
                    st.info("This is the earliest archived run \u2014 there is nothing earlier to compare against.")
                else:
                    prev_id = order_ids[cur_i - 1]
                    prev_df = load_run(prev_id, str(DATA_DIR))

                    def _keyset(d):
                        if d is None or d.empty or "asset_id" not in d or "cve_id" not in d:
                            return set()
                        return set((d["asset_id"].astype(str) + "|" + d["cve_id"].astype(str)).tolist())

                    cur_k, prev_k = _keyset(run_df), _keyset(prev_df)
                    new_k, gone_k, keep_k = cur_k - prev_k, prev_k - cur_k, cur_k & prev_k
                    d = st.columns(3)
                    kpi_card(d[0], "Newly appeared", len(new_k), YELLOW, "#8A6D0E", f"vs {prev_id}")
                    kpi_card(d[1], "Resolved / gone", len(gone_k), GREEN, GREEN_D, f"vs {prev_id}")
                    kpi_card(d[2], "Still open", len(keep_k), GREY, NAVY, f"vs {prev_id}")
                    if new_k:
                        new_view = run_df[(run_df["asset_id"].astype(str) + "|" + run_df["cve_id"].astype(str)).isin(new_k)]
                        nc = [c for c in ["asset_name", "cve_id", "priority", "final_score", "kev_status", "sla_status"] if c in new_view.columns]
                        st.markdown("**Newly appeared in this run**")
                        render_table(new_view[nc], height=220)

            section("Archived Exposure Results", GREEN)
            disp = [c for c in ["asset_name", "application_name", "cve_id", "severity", "cvss_score",
                                "epss_percentile", "kev_status", "network_exposure_level", "ids_alert_count",
                                "privacy_impact_level", "final_score", "priority", "sla_status", "primary_action"]
                    if c in run_df.columns]
            render_table(run_df[disp] if disp else run_df)

            # Re-download this run's Action Plan (worst-first), HMAC-signed
            section("Re-download This Run", GREEN)
            po = {"Critical": 1, "High": 2, "Medium": 3, "Low": 4}
            rr = run_df.copy()
            rr["_k"] = (rr.get("kev_status", "").astype(str).str.lower() == "yes").map({True: 0, False: 1})
            rr["_p"] = rr.get("priority", "").map(po).fillna(5)
            rr["_s"] = pd.to_numeric(rr.get("final_score", 0), errors="coerce").fillna(0)
            rr = rr.sort_values(["_k", "_p", "_s"], ascending=[True, True, False]).drop(columns=["_k", "_p", "_s"])
            hplan = compose_action_plan(rr)
            rid = str(sel["run_id"])
            g = st.columns(3)
            hcsv = to_csv_bytes(hplan)
            with g[0]:
                st.download_button("Action Plan CSV", data=hcsv, file_name=f"action_plan_{rid}.csv",
                                   mime="text/csv", use_container_width=True, key=f"h_csv_{rid}")
                st.code(generate_sha256(hcsv), language="text")
            with g[1]:
                try:
                    hx = to_xlsx_bytes(hplan)
                    st.download_button("Action Plan Excel", data=hx, file_name=f"action_plan_{rid}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True, key=f"h_xlsx_{rid}")
                    st.code(generate_sha256(hx), language="text")
                except Exception as exc:
                    st.warning(f"Excel unavailable: {exc}")
            with g[2]:
                try:
                    hp = to_pdf_bytes(hplan, generated_by=str(sel.get("operator", "system")))
                    st.download_button("Action Plan PDF", data=hp, file_name=f"action_plan_{rid}.pdf",
                                       mime="application/pdf", use_container_width=True, key=f"h_pdf_{rid}")
                    st.code(generate_sha256(hp), language="text")
                except ImportError:
                    st.warning("PDF export needs `reportlab` (add to requirements.txt).")
                except Exception as exc:
                    st.warning(f"PDF unavailable: {exc}")

            with st.expander("Delete this archived run", expanded=False):
                st.caption("Removes this run's stored results and registry entry. This cannot be undone.")
                confirm = st.checkbox("I understand this permanently deletes this archived run.", key=f"del_ok_{rid}")
                if st.button("Delete run", key=f"del_btn_{rid}", disabled=not confirm):
                    if delete_run(rid, str(DATA_DIR)):
                        log_event("Assessment history run deleted", details=f"run_id={rid}", operator=operator())
                        st.success(f"Deleted run {rid}. Refreshing\u2026")
                        st.rerun()
                    else:
                        st.error("Could not delete the run.")

    st.caption(
        "Note: on Cloud Run, local files reset on restart. Set the `CEGP_GCS_BUCKET` environment "
        "variable (and add `google-cloud-storage`) to persist this history to a Cloud Storage bucket "
        "instead \u2014 see the GCP guide, Part 11. Locally it persists on disk."
    )
